# -*- coding: utf-8 -*-
#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
# Name:    LEITURA DOS BIGFILES DO SAMP
# Purpose: MONTA OUTRO ARQUIVO DE SAIDA
#
# Author:      Marcelo Facio Palin
#
# Last Update:  27/02/2015
# Created:     27/02/2015
# Copyright:   (c) mpi 2015
#-------------------------------------------------------------------------------
##########################################
## Outros pacotes Nativos do Python
##########################################
import sys
import os
import getpass ##retorna o Nome do Usuario

##########################################
#### Criptografia
##########################################
try:
    import simplecrypt
except ImportError:
    print "Por favor instale o pacote simplecrypt!"
    sys.exit()
from simplecrypt import encrypt, decrypt

##########################################
#### pyodbc - Conexao com o Access
##########################################
try:
    import pyodbc
except ImportError:
    print "Por favor instale o pacote pyodbc!"
    sys.exit()

##########################################
#### openpyxl - Escreve e Le Excel 2007
##########################################
try:
    import openpyxl
except ImportError:
    print "Por favor instale o pacote openpyxl!"
    sys.exit()

import string
from openpyxl import Workbook
from openpyxl.cell import get_column_letter

##########################################
#### lxml - Leitura de XML
##########################################
try:
    import lxml
except ImportError:
    print "Por favor instale o pacote lxml!"
    sys.exit()

from lxml import etree
from StringIO import StringIO


##########################################
#### BeautifulSoup4 - Leitura de XML
##########################################
try:
    from bs4 import BeautifulSoup
except ImportError:
    print "Por favor instale o pacote BeautifulSoup4!"
    sys.exit()

##########################################
# Outros pacotes Nativos do Python
##########################################
import string
import re ##Expressoes Regulares
import calendar
import datetime
from dateutil.relativedelta import *
import locale
from locale import setlocale, LC_ALL
import math
import operator
import csv
from types import * #Checagem das Variaveis
from xlutils.copy import copy
import shutil, errno
import types
##ProgressBar
##from __future__ import print_function
import sys
import time

##from progressbar import AnimatedMarker, Bar, BouncingBar, Counter, ETA, \
##    FileTransferSpeed, FormatLabel, Percentage, \
##    ProgressBar, ReverseBar, RotatingMarker, \
##    SimpleProgress, Timer, AdaptiveETA, AdaptiveTransferSpeed


## METODO CONTA
def blocks(files, size=65536):

  ##pbar = ProgressBar(widgets=[Percentage(), Bar()], maxval=1).start()
  while True:
    b = files.read(size)
    ##pbar += 1 #Atualiza a Barra de Status
    if not b: break
    yield b

  ##pbar.finish()


###########################################################################
### ( vs1.0 ) - NOVO RELATORIO DE INTELIGENCIA DE MERCADOS - PARA BD
### - LEITURA DOS XLSX E HTMLS nos Diretorios
###########################################################################
def main():

  auth_proxy = 0
  tipoBD = 1 ##Access
  dirBD = ''

  dirini = os.getcwd()

  ##conn.close()
  now = datetime.datetime.now()
  hoje = datetime.datetime.today()
  ##info_System()
  ##print "Diretorio Entrada (onde estao os HTMLs Baixados da CCEE): " + str(dirini)

  ##Verificando se existem os Diretorios:
  dirOutput = "Saida"
  if not os.path.exists(dirOutput):
    os.makedirs(dirOutput)
  dirInput = "Entrada"
  if not os.path.exists(dirInput):
    os.makedirs(dirInput)

  printDebug = 0

  ##@@@@@
  ##@ (1.0)
  ##@ - VARRENDO OS DIRETORIOS EM BUSCA DOS ARQUIVOS
  ## - Definindo a extensao dos arquivos que serao lidos
  ## - Preenchimento da Lista lstPathAndFiles com o Caminho e o Nome dos
  ## arquivos que serao inseridos no BD
  #######################################################################
  extensaoFile =( '.txt')
  lstPathAndFiles = []


  ## LISTA DE ARQUIVOS (SMP)
  lstPathAndFiles = getList_NomeArq_E_Path("Entrada",extensaoFile)
  ##Imprime Log
  strFileNameLog =  "Arquivos_Encontrados.log"
  printLogFilesFound(lstPathAndFiles,strFileNameLog)

  intCount = 1
  ##@@@@@
  ##@ (2.0)
  ##@ - LOOP DE INTERPRETACAO DE CADA ARQUIVO
  #######################################################################
  farqJaProc = open("Arq_Processados.log", "w")
  for i in xrange(0,len(lstPathAndFiles)):

      strFileName = lstPathAndFiles[i][0] ##Dir + Nome
      strDirAndFileName = lstPathAndFiles[i][1]## Apenas Nome

      if(strFileName.find(extensaoFile) != -1): ##Encontrou a Extensao HTM

          ##@@@@@
          ##@ (NEW) - O nome Contem Geral de Consumo por Classe
          if (strFileName.find("ZQM_MP") != -1):

            ## Indica qual arquivo esta lendo
            print "### CADASTRO ###"
            print "(*) Leitura Arquivo: " + strFileName
            print "(*) Dir+Arquivo: " + strDirAndFileName


            with open(strDirAndFileName, "r") as f:
                nTotalLinhas =  sum(bl.count("\n") for bl in blocks(f))
            print "O N. Total de Linhas do ARQUIVO " + strFileName + ": " + str(nTotalLinhas)


            dest_filename = "Amostra_" + strFileName + ".xlsx"
            wb = Workbook()
            ##dest_filename = r'empty_book.xlsx'
            ws = wb.worksheets[0]
            ##Nomeia a Aba
            ws.title = "CADASTRO"

            idRow = 1
            ##LEITURA LINHA A LINHA DO BIG FILE
            with open(strDirAndFileName) as infile:

              for line in infile:
                print line
                lstValores = line.split(';')
                nCols = len(lstValores)
##                UC = lstValores[4] ##0-idx
##                MesRef = lstValores[1]
##                Consumidor = lstValores[14]
##                print Consumidor
##                print "UC = " + str(UC)
                for i in xrange(0,nCols):
                  ws.cell(row = idRow, column = i+1).value = str(lstValores[i]).decode('latin-1').strip()
                idRow = idRow + 1
                if idRow == 400:
                  break

              ws = wb.create_sheet()
              wb.save(filename = dest_filename)

            print "## Terminado  ##\n"
            farqJaProc.write(strDirAndFileName + "\n")


  farqJaProc.close()

  #######################################################################
  ## (FINAL )
  ## -corrige os valores que foram inseridos como NONE para Zero
  #######################################################################
  ## CORRIGE OS VALORES COLOCANDO ZERO NOS CAMPOS VAZIOS
  ## Fazendo a correcao por Ano para caber os registros na Memoria
  ##strIDAnoToCorr = str(now.year)
  ##update_TabPubCCEE_WithZeros(conn,strIDAnoToCorr)
  ##print "Apenas UPDATE de " + strIDAnoToCorr + " - saindo!"

##  conn.close()

######################### FINAL DO MAIN ########################################


################################################################################
## METODO @2014 - INSERE Geracao Final (MWh) Relatorio Publico CCEE
## Geracao Final (MWh) por Semana Patamar
################################################################################
def Insere_GeracaoFinal_CentroGrav_MWh(conn,lstRowsTab,dtRef_RelatPublico,printDebug = False):

    ## CHECK - Keys
    sqlcheck_Patamar = "SELECT idAgente, Mes, Ano FROM Agentes_Patamar_PublicoCCEE WHERE ((idAgente=[ID_AGENTE]) AND (Mes=[MES]) AND (Ano=[ANO]) AND (Patamar='[PATAMAR]') )"
    ## INSERINDO Keys
    sqlins_Patamar = "INSERT INTO Agentes_Patamar_PublicoCCEE(idAgente, Mes, Ano, Patamar) VALUES ([ID_AGENTE], [MES], [ANO],'[PATAMAR]')"
    ## UPDATE - valor
    sqlupd_Patamar = "UPDATE Agentes_Patamar_PublicoCCEE SET [VALORES] WHERE ((idAgente=[ID_AGENTE]) AND (Mes=[MES]) AND (Ano=[ANO]) AND (Patamar='[PATAMAR]') )"


    ## CHECK - Keys
    sqlcheck_Mensal = "SELECT idAgente, Mes, Ano FROM Agentes_Mensal_PublicoCCEE WHERE ((idAgente=[ID_AGENTE]) AND (Mes=[MES]) AND (Ano=[ANO]))"
    ## INSERINDO Keys
    sqlins_Mensal = "INSERT INTO Agentes_Mensal_PublicoCCEE(idAgente, Mes, Ano) VALUES ([ID_AGENTE], [MES], [ANO])"
    ## UPDATE - valor
    sqlupd_Mensal = "UPDATE Agentes_Mensal_PublicoCCEE SET [VALORES] WHERE ((idAgente=[ID_AGENTE]) AND (Mes=[MES]) AND (Ano=[ANO]))"

    ##@@@@@
    ##@ (1.0) - DEBUG - IMPRIME O QUE SERA INSERIDO
##    if printDebug == True:
##      strFileNameLog =  "Saida/" + str(dtRef_RelatPublico.year) + "_" + str(dtRef_RelatPublico.month).zfill(2) + "_linhas_MedGeracao" + ".txt"
##      f = open(strFileNameLog, 'w')
##      for i in xrange(0,len(lstRowsTab)):
##        row = lstRowsTab[i]
##        val_PatLeve = row[2].replace("\xa0","").strip()
##        val_PatMedio = row[3].replace("\xa0","").strip()
##        val_PatPesado = row[4].replace("\xa0","").strip()
##
##        if (not val_PatLeve): ##Se Nao Ha String (Vazia) - coloque 0.0
##          val_PatLeve = '0'
##
##        if (not val_PatMedio):
##          val_PatMedio = '0'
##
##        if (not val_PatPesado):
##          val_PatPesado = '0'
##
##        val_PatLeve = val_PatLeve.replace(",",".")
##        val_PatMedio = val_PatMedio.replace(",",".")
##        val_PatPesado = val_PatPesado.replace(",",".")
##        valTotalCons = float(val_PatLeve) + float(val_PatMedio) + float(val_PatPesado)
##        strCol3_Valor = str(valTotalCons)
##        ##Para imprimir o CSV, substitui PONTO por VIRGULA
##        strCol3_Valor = strCol3_Valor.replace(".",",")
##        f.write(row[0] + ";" + row[1] + ";" + strCol3_Valor + "\n")
##      f.close()

    dia = dtRef_RelatPublico.day
    mes = dtRef_RelatPublico.month
    ano = dtRef_RelatPublico.year

    if printDebug == True:
      print "##########################################"
      print " INSERINDO Geracao (MWh) Centro Grav Gpj (Html) (" + str(dtRef_RelatPublico.year) + "/" + str(dtRef_RelatPublico.month).zfill(2) + ")"
      print "##########################################"

    ##@@@@@
    ##@ (2.0) - LOOP - por TODAS as Linhas da TABELA
    intTotalReg = len(lstRowsTab)
    count = 0

    for row in lstRowsTab:
        count = count + 1
        if ( (count % 100) == 0):
          porcent = (float(count)/float(intTotalReg))
          print ("inserido %10.2f porc" % (float(porcent)*100.) )

        strSiglaAgente = ""
        strRazSocAgente = ""
        try:
          strSiglaAgente = str(row[0]).strip()
          strRazSocAgente = str(row[1]).strip() ##strip() Remove espacos do Inicio e Fim
        except:
            if printDebug == True:
              print "Linha em Branco - Pulando!"
            continue

        print "Estou no Registro: " + str(row) + " do Agente " + strSiglaAgente

        ##@@@@@
        ##@ (2.1) - Check Agente - pegando o ID do BD Local
        ## - se Nao Existir Insere na Tabela Agentes e Agentes_Sinonimos
        ## e MARCA que foi inserido do PublicoCCEE LIDO DO MES/ANO na origem_agente
        ## e MARCA que o CodCliq e' TEMPORARIO
##        if printDebug == 1:
##          print "LINHA MED GERACAO: " + str(count)
##          print "Pegando o ID do Agente: " + strSiglaAgente

        idAgente = CheckAgente_FromTabSinonimos(conn, strSiglaAgente, dtRef_RelatPublico, True)
        if printDebug == 1:
          print "Peguei o ID Agente: " + strSiglaAgente + " = " + str(idAgente)

        ##[(1)CHECK na Tabela: Agentes_Mensal_PublicoCCEE]
        sql = sqlcheck_Mensal.replace("[ID_AGENTE]",str(idAgente))
        sql = sql.replace("[MES]", str(mes))
        sql = sql.replace("[ANO]", str(ano))
        cur = conn.cursor()

        cur.execute(sql)
        result = cur.fetchone()
        if (result == None):
          ##[(1) INSERT na Tabela: Agentes_Mensal_PublicoCCEE]
          sql = sqlins_Mensal.replace("[ID_AGENTE]",str(idAgente))
          sql = sql.replace("[MES]", str(mes))
          sql = sql.replace("[ANO]", str(ano))
          cur = conn.cursor()
          try:
              if printDebug == 1:
                print "ID Agente: " + strSiglaAgente + " = " + str(idAgente)
                print "SQL: " + sql
              cur.execute(sql)
          except:
              print "ERRO INSERT na keys da Tab Agentes_Mensal_PublicoCCEE Agente: " + strSiglaAgente + " = " + str(idAgente) + " Mes: " + str(mes) + " Ano: " + str(ano)
        else:
          pass
        cur.close()

        ##LOOP PARA OS 3 PATAMARES (L,M,P)
        lstPatamares = ['L','M','P']

        for i in xrange(0,3):
          ##[(2) CHECK na Tabela: Agentes_Patamar_PublicoCCEE]
          sql = sqlcheck_Patamar.replace("[ID_AGENTE]",str(idAgente))
          sql = sql.replace("[MES]", str(mes))
          sql = sql.replace("[ANO]", str(ano))
          sql = sql.replace("[PATAMAR]", str(lstPatamares[i]))
          cur = conn.cursor()

          cur.execute(sql)
          result = cur.fetchone()
          if (result == None):
            ##[(2) INSERT na Tabela: Agente_SemanaPatamar_PublicoCCEE]
            sql = sqlins_Patamar.replace("[ID_AGENTE]",str(idAgente))
            sql = sql.replace("[MES]", str(mes))
            sql = sql.replace("[ANO]", str(ano))
            sql = sql.replace("[PATAMAR]", str(lstPatamares[i]))
            cur = conn.cursor()
            try:
                cur.execute(sql)
            except:
                print "ERRO na keys de Agentes_Mensal_PublicoCCEE: " + str(idAgente) + " Mes: " + str(mes) + " Ano: " + str(ano)
          else:
            pass
          cur.close()


        lstValoresPatamares = [0,0,0]
        ##LISTA DE VALORES A SEREM INSERIDO - UPDATE
        val_PatLeve = row[2].replace("\xa0","").strip() ##Remove espacos e \xa0
        val_PatMedio = row[3].replace("\xa0","").strip() ##Remove espacos
        val_PatPesado = row[4].replace("\xa0","").strip() ##Remove espacos

        if (not val_PatLeve): #Se Nao Ha' String (Vazia) - coloque 0.0
          val_PatLeve = '0'

        if (not val_PatMedio):
          val_PatMedio = '0'

        if (not val_PatPesado):
          val_PatPesado = '0'

        val_PatLeve = val_PatLeve.replace(",",".")
        val_PatMedio = val_PatMedio.replace(",",".")
        val_PatPesado = val_PatPesado.replace(",",".")
        lstValoresPatamares[0] = val_PatLeve
        lstValoresPatamares[1] = val_PatMedio
        lstValoresPatamares[2] = val_PatPesado

        valTotalCons = float(val_PatLeve) + float(val_PatMedio) + float(val_PatPesado)

        if (not valTotalCons): ##Se Nao Ha' String (Vazia) - coloque 0.0
          valTotalCons = '0'

        lst_NomeCampo_Valor = []
        ##@@ (1)
        lst_NomeCampo_Valor.append("ger_gpj_cg_mwh = " + str(valTotalCons))
        now = datetime.datetime.now()

        ##@@ (2)
        lst_NomeCampo_Valor.append("origem_geracao_gpj_cg = '" + "PUBLICO CCEE" + str(dtRef_RelatPublico.year) + "/" + str(dtRef_RelatPublico.month).zfill(2)  + "'" )

        now = datetime.datetime.now()
        strNomeUsuario = getpass.getuser()
        ##@@ (3)
        ##lst_NomeCampo_Valor.append("QuemAtualizou = '" + strNomeUsuario + "'")
        ##@@ (4)
        lst_NomeCampo_Valor.append("DataAtualizacao_Geracao = '" + now.strftime("%d/%m/%Y %H:%M:%S") + "'")

        valores = ",".join(lst_NomeCampo_Valor)
        sql = sqlupd_Mensal.replace("[ID_AGENTE]",str(idAgente))
        sql = sql.replace("[MES]", str(mes))
        sql = sql.replace("[ANO]", str(ano))
        sql = sql.replace("[VALORES]", valores)
        cur = conn.cursor()
        ##print sql
        try:
            cur.execute(sql)
        except:
            print "Erro no UPDATE MED GERACAO idAgente = " + str(idAgente)
        conn.commit()

        ########################################################################
        ##=============== PARTE II ==========================
        ## INSERE OS VALORES DE GERACAO POR PATAMAR EM OUTRA TABELA
        ########################################################################
        for i in xrange(0,3):
          lst_NomeCampo_Valor = []
          ##@@ (1) Relatorio Individual de medicao
          lst_NomeCampo_Valor.append("ger_gpj_cg_mwh = " + str(lstValoresPatamares[i])) ##Geracao Agente no Centro de Gravidade (MWh)
          now = datetime.datetime.now()

          ##@@ (2)
          lst_NomeCampo_Valor.append("origem_geracao_gpj_cg = '" + "PUBLICO CCEE" + str(dtRef_RelatPublico.year) + "/" + str(dtRef_RelatPublico.month).zfill(2)  + "'" )

          now = datetime.datetime.now()

          ##@@ (3)
          lst_NomeCampo_Valor.append("DataAtualizacao_Geracao = '" + now.strftime("%d/%m/%Y %H:%M:%S") + "'")

          valores = ",".join(lst_NomeCampo_Valor)
          sql = sqlupd_Patamar.replace("[ID_AGENTE]",str(idAgente))
          sql = sql.replace("[MES]", str(mes))
          sql = sql.replace("[ANO]", str(ano))
          sql = sql.replace("[PATAMAR]", str(lstPatamares[i]))
          sql = sql.replace("[VALORES]", valores)
          cur = conn.cursor()
          ##print sql
          try:
              cur.execute(sql)
          except:
              print "Erro no UPDATE MED GERACAO idAgente = " + str(idAgente)
          conn.commit()




##------------------------------------------------------------------------------






################################################################################
## METODO I @2014 - INSERE CONTRATOS DE COMPRA NO BD - por XLSX
################################################################################
def Insere_Consumo_ClasseAgente_BD(conn, dic_ClasseAgente_ConsumoMWh, dtRef ,printDebug = False):
    ## CHECK - Keys
    sql_check = "SELECT idTipo FROM Carga_ClasseAgente_Mensal WHERE ( (idTipo=[ID_TIPO]) AND (Mes=[MES]) AND (Ano=[ANO]) )"
    ## INSERINDO Keys
    sql_ins = "INSERT INTO Carga_ClasseAgente_Mensal(idTipo, Mes, Ano) VALUES ([ID_TIPO], [MES], [ANO])"
    ## UPDATE - valor
    sql_upd = "UPDATE Carga_ClasseAgente_Mensal SET [VALORES] WHERE ( (idTipo=[ID_TIPO]) AND (Mes=[MES]) AND (Ano=[ANO]) )"

    ##@@@@@
    ##@  -- MONTANDO O DICIONARIO DE TIPOS DE CLASSE DE AGENTES
    ##@  -- PARA PODERMOS INSERIR NO BD NAO O TEXTO, MAS SIM O ID DO TIPO
    sql_tipos_classe_agentes = "SELECT IdTipo, Tipo FROM Agentes_Tipo"
    cur = conn.cursor()
    cur.execute(sql_tipos_classe_agentes)
    rows = cur.fetchall()
    ##@@@@@
    ##@ (1.0)
    ##@ mapeia Autoprodutor = 1, Comercializador = 2, Distribuidor = 4
    dic_strTipoClasse_ID = {}
    if (rows != None):
      for row in rows:
          ## ONE KEY -> MULTIPLE VALUES
          ##dic_strTipoUsina_ID.setdefault(row.AbrevTipoUsina, []).append(row.idTipoUsina)
          ##Aqui nao e' o Caso, 1 Key - 1 Valor
          dic_strTipoClasse_ID[row.Tipo] = row.IdTipo
          ##print row.Tipo, row.IdTipo
    else:
      print "## PROBLEMA AO CONSULTAR A TABELA Agentes_Tipo ##"
      pass ##Pula
    cur.close()

    dia = dtRef.day
    mes = dtRef.month
    ano = dtRef.year

    print "##########################################"
    print " INSERINDO CONSUMO POR CLASSE DE (" + str(dtRef.year) + "/" + str(dtRef.month).zfill(2) + ")"
    print "##########################################"

    for key, valor_consumo_mwh in dic_ClasseAgente_ConsumoMWh.items():
      if key in dic_strTipoClasse_ID:
        idTipoClasse = dic_strTipoClasse_ID[key]
      else:
        print "ERRO, NAO ENCONTRADO O TIPO " + key + "NA TABELA Agentes_Tipo"
        exit

      ##CHECK KEYS - idTipo, Ano, Mes
      sql = sql_check.replace("[ID_TIPO]",str(idTipoClasse))
      sql = sql.replace("[MES]", str(mes))
      sql = sql.replace("[ANO]", str(ano))
      cur = conn.cursor()
      cur.execute(sql)
      result = cur.fetchone()
      ##@[INSERT]
      if (result == None):
        sql = sql_ins.replace("[ID_TIPO]",str(idTipoClasse))
        sql = sql.replace("[MES]", str(mes))
        sql = sql.replace("[ANO]", str(ano))
        cur = conn.cursor()
        try:
            cur.execute(sql)
        except:
            print "Nao consegui inserir as keys: " + str(idTipoClasse) + " Mes: " + str(mes) + " Ano: " + str(ano)
      else:
        pass
      cur.close()

      if (not valor_consumo_mwh): ##Se Nao Ha' String (Vazia) - coloque 0.0
        valor_consumo_mwh = '0'

      lst_NomeCampo_Valor = []
      ##@@ (1)
      lst_NomeCampo_Valor.append("ConsumoClasse_MWh = " + str(valor_consumo_mwh))
      now = datetime.datetime.now()
      strNomeUsuario = getpass.getuser()
      ##ADICIONANDO 2 CAMPOS
      lst_NomeCampo_Valor.append("QuemInseriu = '" + strNomeUsuario + "'")
      lst_NomeCampo_Valor.append("DataInsercao = '" + now.strftime("%d/%m/%Y %H:%M:%S") + "'")

      valores = ",".join(lst_NomeCampo_Valor)
      sql = sql_upd.replace("[ID_TIPO]",str(idTipoClasse))
      sql = sql.replace("[MES]", str(mes))
      sql = sql.replace("[ANO]", str(ano))
      sql = sql.replace("[VALORES]", valores)
      cur = conn.cursor()
      try:
          cur.execute(sql)
      except:
          print "Erro no UPDATE Carga_ClasseAgente_Mensal!"
      conn.commit()
##------------------------------------------------------------------------------




################################################################################
## METODO - EXTRAI CONTEUDO DE ARQUIVO - DADO EXPRESSAO REGULAR
################################################################################
## http://stackoverflow.com/questions/7098530/repeatedly-extract-a-line-between-two-delimiters-in-a-text-file-python
def getBlockText_From_String(strTexto, strSubDir, strExpressRegular, strFileSaida, printDebug = False):
  for strConteudo in re.findall(strExpressRegular, strTexto, re.S):
    strBloco = strConteudo
    if printDebug == 1:
      try:
        f = open(os.getcwd() + '/' + strSubDir + '/' + strFileSaida, "w")
        try:
          f.writelines(strConteudo)
        finally:
          f.close()
      except IOError:
        pass

    return strBloco
##------------------------------------------------------------------------------



################################################################################
## METODO - EXTRAI CONTEUDO DE ARQUIVO - DADO EXPRESSAO REGULAR
################################################################################
## http://stackoverflow.com/questions/7098530/repeatedly-extract-a-line-between-two-delimiters-in-a-text-file-python
def getBlockText_From_File(strFName, strSubDir, strExpressRegular, strFileSaida, printDebug = False):

    with open(os.getcwd() + '/' + strSubDir + '/' + strFName) as fp:
      ##@@ PROCURA O CONTEUDO DADO PELA EXPRESSAO REGULAR
      ## PASSADA COMO PARAMETRO
      for strConteudo in re.findall(strExpressRegular, fp.read(), re.S):
        strBloco = strConteudo
        if printDebug == 1:
          try:
            f = open(os.getcwd() + '/' + strSubDir + '/' + strFileSaida, "w")
            try:
              f.writelines(strConteudo)
            finally:
              f.close()
          except IOError:
            pass
    return strBloco
##------------------------------------------------------------------------------



################################################################################
## METODO - GRAB TABELA
## Relatorio Mensal de Consumo por Classe de Agente
## Esse HTML e' um lixo tem 23 tabelas para representar apenas 1 tabela
## Solucao: peguei o texto VISIVEL e utilizei expressoes regulares para pegar
## o conteudo.
################################################################################

def grab_Tabela_Consum_ClasseAgente(strConteudoHTML, printDebug = False):
  ##@@@@@
  ##@ (1.0)
  ##@ - BeautifulSoup - interpreta o HTML
  soup = BeautifulSoup(strConteudoHTML)

  ##BeautifulSoup LISTA DE TABELAS DA PAGINA
  lstTables = soup.find_all('table')
  if printDebug == True:
    print "Existem " + str(len(lstTables)) + " tabelas na pagina"

  strBlocoTXT_Visivel = soup.body.getText()
  if printDebug == True:
    writeFile("Saida/HTML_Consumo_ClasseAgente.txt",strBlocoTXT_Visivel)

  ##@@ (2.0) Procura pela data dos Valores Ex: 2014/01
  ##(1o) Find
  ##dd/mm/aaaa - PADRAO DE DATA PROCURADO
  ##expRegDATA = re.compile(r'([0-9]|[0,1,2][0-9]|3[0,1])/(0[0-9]|[\d]|1[0,1,2])/(\d{4})')
  ##aaaa/mm
  strExpReg = re.compile(r'(\d{4}/\d{2})')
  ##Search de re - procupara a 1a ocorrencia
  match = re.search(strExpReg, strBlocoTXT_Visivel)
  if match:
    strAnoMes = match.group(0)

  strAno = ''
  strMes = ''
  if strAnoMes != '':
    strAno, strMes = strAnoMes.split('/')



  ##@@ MERCADO DE ENERGIA TOTAL - SISTEMA.DAT
  strEReg = 'Consumo Total Registrado \(MWh\)(.*?)Total Geral'
  ##Ultimo parametro imprime arquivo com o Bloco Extraido
  strFileSaida = 'strBloco_TabConsumo_Por_Classe.txt'
  strBlocoTabela = getBlockText_From_String(strBlocoTXT_Visivel,"Saida",strEReg,strFileSaida,True)


  ## Armazena TODAS as linhas do Arquivo em uma LISTA
  lstLines = strBlocoTabela.split('\n')
  if printDebug == True:
    print "Total de Linhas a serem lidas = " + str(len(lstLines))

  lstFilterLines = []
  lstFilterLines = filter(None, lstLines)

  ## Remove os espacos das linhas "teoricamente em branco" e entao
  ## testa se a funcao temporaria lambda resultar numa lista de len = 0
  ## a linha e' para ser FILTRADA (removida)
  ##r = filter(lambda x: len(x.split()) != 0, ['1  ', '   ', '   2'])
  ##Filtrando Novamente a lstFilterLines para remover as linhas (elementos)
  ## que eram linhas apenas com espacoes em branco
  lstFilterLines = filter(lambda x: len(x.split()) != 0, lstFilterLines)

  if printDebug == True:
    f = open("Saida/LinhasFiltradas_ConsumoPorClasse.txt", "w")
    for i in xrange(0,len(lstFilterLines)):
      f.write(lstFilterLines[i]+"\n")
    f.close()

  ## DEPOIS DE TUDO TEREMOS:
  ##Autoprodutor
  ##2.476.139,464
  ##Consumidor Especial
  ##1.432.489,422
  ##Consumidor Livre
  ##7.200.283,051
  ##Distribuidor
  ##36.241.902,629
  ##Gerador
  ##743.504,650
  ##Importador
  ##0,000

  dic_ClasseAgente_ConsumoMWh = {}
  strClasse = ''
  strNumber = ''
  for line in lstFilterLines:
    if line != '': ##Codigo de Seguranca
      strTxt = line
      strTxt = strTxt.replace(".","")
      strTxt = strTxt.replace(",",".")
      if is_number(strTxt) == True:
        strNumber = strTxt
        dic_ClasseAgente_ConsumoMWh[strClasse] = strNumber
      else:
        strClasse = strTxt

  return dic_ClasseAgente_ConsumoMWh,strAno, strMes
##------------------------------------------------------------------------------





################################################################################
## METODO I @2014 - INSERE CONTRATOS DE COMPRA NO BD - por XLSX
################################################################################
def Insere_GF_XLSX_MONTADO_ANTES_(conn,lstRowsTabContCompra,dtRef_RelatPublico,printDebug = False):
    ## CHECK - Keys
    sqlcheck = "SELECT idAgente, Mes, Ano FROM Agentes_Mensal_PublicoCCEE WHERE ((idAgente=[ID_AGENTE]) AND (Mes=[MES]) AND (mes=[MES]) AND (Ano=[ANO]))"
    ## INSERINDO Keys
    sqlins = "INSERT INTO Agentes_Mensal_PublicoCCEE(idAgente, Mes, Ano) VALUES ([ID_AGENTE], [MES], [ANO])"
    ## UPDATE - valor
    sqlupd = "UPDATE Agentes_Mensal_PublicoCCEE SET [VALORES] WHERE ((idAgente=[ID_AGENTE]) AND (Mes=[MES]) AND (mes=[MES]) AND (Ano=[ANO]))"

    if printDebug == True:
      strFileNameLog =  "Saida/" + str(dtRef_RelatPublico.year) + "_" + str(dtRef_RelatPublico.month).zfill(2) + "Contratos_Compra_XLSX.txt"
      f = open(strFileNameLog, 'w')
      for i in xrange(0,len(lstRowsTabContCompra)):
        row = lstRowsTabContCompra[i]
        strCol3_Valor = row[2]
        ##Para imprimir o CSV, substitui PONTO por VIRGULA
        strCol3_Valor = strCol3_Valor.replace(".",",")
        f.write(row[0] + ";" + row[1] + ";" + strCol3_Valor + "\n")
      f.close()

    dia = dtRef_RelatPublico.day
    mes = dtRef_RelatPublico.month
    ano = dtRef_RelatPublico.year

    print "##########################################"
    print " INSERINDO CONTRATOS DE COMPRA XLS (" + str(dtRef_RelatPublico.year) + "/" + str(dtRef_RelatPublico.month).zfill(2) + ")"
    print "##########################################"

    intTotalReg = len(lstRowsTabContCompra)
    count = 0
    for row in lstRowsTabContCompra:
        count = count + 1
        if ( (count % 100) == 0):
          porcent = (float(count)/float(intTotalReg))
          print ("inserido %10.2f porc" % (float(porcent)*100.) )

        strSiglaAgente = ""
        strRazSocAgente = ""
        try:
          strSiglaAgente = str(row[0]).strip()
          strRazSocAgente = str(row[1]).strip() ##strip() Remove espacos do Inicio e Fim
        except:
            if printDebug == True:
              print "Linha em Branco - Pulando!"
            continue

        ##@@@@@
        ##@ (2.1) - Check Agente - pegando o ID do BD Local
        ## - se Nao Existir Insere na Tabela Agentes e Agentes_Sinonimos
        ## e MARCA que foi inserido do PublicoCCEE LIDO DO MES/ANO na origem_agente
        ## e MARCA que o CodCliq e' TEMPORARIO
        idAgente = CheckAgente_FromTabSinonimos(conn, strSiglaAgente, dtRef_RelatPublico, True)

        ## Check se as Chaves Existe
        sql = sqlcheck.replace("[ID_AGENTE]",str(idAgente))
        sql = sql.replace("[MES]", str(mes))
        sql = sql.replace("[ANO]", str(ano))
        cur = conn.cursor()
        cur.execute(sql)
        result = cur.fetchone()
        ##@[INSERT]
        if (result == None):
          ##@ INSERT DADOS EM Agentes_Mensal_PublicoCCEE
          sql = sqlins.replace("[ID_AGENTE]",str(idAgente))
          sql = sql.replace("[MES]", str(mes))
          sql = sql.replace("[ANO]", str(ano))
          cur = conn.cursor()
          try:
              cur.execute(sql)
          except:
              print "Nao consegui inserir as keys: " + str(idAgente) + " Mes: " + str(mes) + " Ano: " + str(ano)
        else:
          pass
        cur.close()


        ##LISTA DE VALORES A SEREM INSERIDO - UPDATE
        valCompra = row[2].strip() ##Remove espacos

        if (not valCompra): ##Se Nao Ha' String (Vazia) - coloque 0.0
          valCompra = '0'

        valCompra = valCompra.replace(",",".") ##No Access inserimos com Decimal separado por PONTO
        lst_NomeCampo_Valor = []
        ##@@ (1)
        lst_NomeCampo_Valor.append("contrato_compra_mwh = " + str(valCompra))
        now = datetime.datetime.now()

        ##@@ (2)
        lst_NomeCampo_Valor.append("origem_contrato_compra = '" + "PUBLICO CCEE" + "'" )

        now = datetime.datetime.now()
        strNomeUsuario = getpass.getuser()
        ##@@ (3)
        ##lst_NomeCampo_Valor.append("QuemAtualizou = '" + strNomeUsuario + "'")

        ##@@ (4)
        lst_NomeCampo_Valor.append("DataAtualizacao_GF = '" + now.strftime("%d/%m/%Y %H:%M:%S") + "'")

        valores = ",".join(lst_NomeCampo_Valor)
        sql = sqlupd.replace("[ID_AGENTE]",str(idAgente))
        sql = sql.replace("[MES]", str(mes))
        sql = sql.replace("[ANO]", str(ano))
        sql = sql.replace("[VALORES]", valores)
        cur = conn.cursor()
        try:
            cur.execute(sql)
        except:
            print "Erro no UPDATE GFisica XLSX do Registro!"
        conn.commit()
##------------------------------------------------------------------------------







################################################################################
## METODO I @2014 - INSERE CONTRATOS DE COMPRA NO BD - por XLSX
################################################################################
def Insere_ContratosCompra_FromXLS_BD(conn,lstRowsTabContCompra,dtRef_RelatPublico,printDebug = False):
    ## CHECK - Keys
    sqlcheck = "SELECT idAgente, Mes, Ano FROM Agentes_Mensal_PublicoCCEE WHERE ((idAgente=[ID_AGENTE]) AND (Mes=[MES]) AND (mes=[MES]) AND (Ano=[ANO]))"
    ## INSERINDO Keys
    sqlins = "INSERT INTO Agentes_Mensal_PublicoCCEE(idAgente, Mes, Ano) VALUES ([ID_AGENTE], [MES], [ANO])"
    ## UPDATE - valor
    sqlupd = "UPDATE Agentes_Mensal_PublicoCCEE SET [VALORES] WHERE ((idAgente=[ID_AGENTE]) AND (Mes=[MES]) AND (mes=[MES]) AND (Ano=[ANO]))"

    if printDebug == True:
      strFileNameLog =  "Saida/" + str(dtRef_RelatPublico.year) + "_" + str(dtRef_RelatPublico.month).zfill(2) + "Contratos_Compra_XLSX.txt"
      f = open(strFileNameLog, 'w')
      for i in xrange(0,len(lstRowsTabContCompra)):
        row = lstRowsTabContCompra[i]
        strCol3_Valor = row[2]
        ##Para imprimir o CSV, substitui PONTO por VIRGULA
        strCol3_Valor = strCol3_Valor.replace(".",",")
        f.write(row[0] + ";" + row[1] + ";" + strCol3_Valor + "\n")
      f.close()

    dia = dtRef_RelatPublico.day
    mes = dtRef_RelatPublico.month
    ano = dtRef_RelatPublico.year

    print "##########################################"
    print " INSERINDO CONTRATOS DE COMPRA XLS (" + str(dtRef_RelatPublico.year) + "/" + str(dtRef_RelatPublico.month).zfill(2) + ")"
    print "##########################################"

    intTotalReg = len(lstRowsTabContCompra)
    count = 0
    for row in lstRowsTabContCompra:
        count = count + 1
        if ( (count % 100) == 0):
          porcent = (float(count)/float(intTotalReg))
          print ("inserido %10.2f porc" % (float(porcent)*100.) )

        strSiglaAgente = ""
        strRazSocAgente = ""
        try:
          strSiglaAgente = str(row[0]).strip()
          strRazSocAgente = str(row[1]).strip() ##strip() Remove espacos do Inicio e Fim
        except:
            if printDebug == True:
              print "Linha em Branco - Pulando!"
            continue

        ##@@@@@
        ##@ (2.1) - Check Agente - pegando o ID do BD Local
        ## - se Nao Existir Insere na Tabela Agentes e Agentes_Sinonimos
        ## e MARCA que foi inserido do PublicoCCEE LIDO DO MES/ANO na origem_agente
        ## e MARCA que o CodCliq e' TEMPORARIO
        idAgente = CheckAgente_FromTabSinonimos(conn, strSiglaAgente, dtRef_RelatPublico, True)

        ## Check se as Chaves Existe
        sql = sqlcheck.replace("[ID_AGENTE]",str(idAgente))
        sql = sql.replace("[MES]", str(mes))
        sql = sql.replace("[ANO]", str(ano))
        cur = conn.cursor()
        cur.execute(sql)
        result = cur.fetchone()
        ##@[INSERT]
        if (result == None):
          ##@ INSERT DADOS EM Agentes_Mensal_PublicoCCEE
          sql = sqlins.replace("[ID_AGENTE]",str(idAgente))
          sql = sql.replace("[MES]", str(mes))
          sql = sql.replace("[ANO]", str(ano))
          cur = conn.cursor()
          try:
              cur.execute(sql)
          except:
              print "Nao consegui inserir as keys: " + str(idAgente) + " Mes: " + str(mes) + " Ano: " + str(ano)
        else:
          pass
        cur.close()


        ##LISTA DE VALORES A SEREM INSERIDO - UPDATE
        valCompra = row[2].strip() ##Remove espacos

        if (not valCompra): ##Se Nao Ha' String (Vazia) - coloque 0.0
          valCompra = '0'

        valCompra = valCompra.replace(",",".") ##No Access inserimos com Decimal separado por PONTO
        lst_NomeCampo_Valor = []
        ##@@ (1)
        lst_NomeCampo_Valor.append("contrato_compra_mwh = " + str(valCompra))
        now = datetime.datetime.now()

        ##@@ (2)
        lst_NomeCampo_Valor.append("origem_contrato_compra = '" + "PUBLICO CCEE" + "'" )

        now = datetime.datetime.now()
        strNomeUsuario = getpass.getuser()
        ##@@ (3)
        ##lst_NomeCampo_Valor.append("QuemAtualizou = '" + strNomeUsuario + "'")

        ##@@ (4)
        lst_NomeCampo_Valor.append("DataAtualizacao_CCompra = '" + now.strftime("%d/%m/%Y %H:%M:%S") + "'")

        valores = ",".join(lst_NomeCampo_Valor)
        sql = sqlupd.replace("[ID_AGENTE]",str(idAgente))
        sql = sql.replace("[MES]", str(mes))
        sql = sql.replace("[ANO]", str(ano))
        sql = sql.replace("[VALORES]", valores)
        cur = conn.cursor()
        try:
            cur.execute(sql)
        except:
            print "Erro no UPDATE CCompra XLSX do Registro!"
        conn.commit()
##------------------------------------------------------------------------------




################################################################################
## METODO - INSERE CONTRATOS DE COMPRA NO BD
##
## EXEMPLO: DADO EXCEL COM NOME AGENTE, RAZ SOCIAL E VALOR
################################################################################
def Insere_ContratosVenda_FromXLS_BD(conn,lstRowsTabContVenda,dtRef_RelatPublico,printDebug = False):
    ## CHECK - Keys
    sqlcheck = "SELECT idAgente, Mes, Ano FROM Agentes_Mensal_PublicoCCEE WHERE ((idAgente=[ID_AGENTE]) AND (Mes=[MES]) AND (mes=[MES]) AND (Ano=[ANO]))"
    ## INSERINDO Keys
    sqlins = "INSERT INTO Agentes_Mensal_PublicoCCEE(idAgente, Mes, Ano) VALUES ([ID_AGENTE], [MES], [ANO])"
    ## UPDATE - valor
    sqlupd = "UPDATE Agentes_Mensal_PublicoCCEE SET [VALORES] WHERE ((idAgente=[ID_AGENTE]) AND (Mes=[MES]) AND (mes=[MES]) AND (Ano=[ANO]))"

    if printDebug == True:
      strFileNameLog =   str(dtRef_RelatPublico.year) + "_" + str(dtRef_RelatPublico.month).zfill(2) + "_CVenda_XLSX_" + ".csv"
      f = open(strFileNameLog, 'w')
      for i in xrange(0,len(lstRowsTabContVenda)):
        row = lstRowsTabContVenda[i]
        strCol3_Valor = row[2]
        ##Para imprimir o CSV, substitui PONTO por VIRGULA
        strCol3_Valor = strCol3_Valor.replace(".",",")
        f.write(row[0] + ";" + row[1] + ";" + strCol3_Valor + "\n")
      f.close()

    dia = dtRef_RelatPublico.day
    mes = dtRef_RelatPublico.month
    ano = dtRef_RelatPublico.year

    print "##########################################"
    print " INSERINDO CONTRATOS DE VENDA XLS ("  + str(dtRef_RelatPublico.year) + "/" + str(dtRef_RelatPublico.month).zfill(2) + ")"
    print "##########################################"

    intTotalReg = len(lstRowsTabContVenda)
    count = 0
    ## ja foi removido da Tabela de Venda
    for row in lstRowsTabContVenda:
        count = count + 1

        if ( (count % 100) == 0):
          porcent = (float(count)/float(intTotalReg))
          print ("inserido %10.2f porc" % (float(porcent)*100.) )

        strSiglaAgente = ""
        strRazSocAgente = ""
        try:
          ##print "Nome SiglaAgente: " + row[0]
          strSiglaAgente = str(row[0]).strip()
          strRazSocAgente = str(row[1]).strip() ##strip() Remove espacos do Inicio e Fim
        except:
            print "Linha em Branco - Pulando!"
            continue

        idAgente = CheckAgente_FromTabSinonimos(conn, strSiglaAgente, dtRef_RelatPublico, True)

        ##print "\n\n" + "##########"
        ##print "SiglaAgente: " + strSiglaAgente +  " ID = " + str(idAgente)
        # Check se as Chaves Existe
        sql = sqlcheck.replace("[ID_AGENTE]",str(idAgente))
        sql = sql.replace("[MES]", str(mes))
        sql = sql.replace("[ANO]", str(ano))
        cur = conn.cursor()

        cur.execute(sql)
        result = cur.fetchone()  # busca o resultado da consulta
        #NAO ENCONTRO O CONJUNTO DE CHAVES - INSERE O CONJUNTO de KEYS
        if (result == None):
          #print "Inserindo Registro"
          sql = sqlins.replace("[ID_AGENTE]",str(idAgente))
          sql = sql.replace("[MES]", str(mes))
          sql = sql.replace("[ANO]", str(ano))
          cur = conn.cursor()
          try:
              cur.execute(sql)
          except:
              print "Nao consegui inserir as keys: " + str(idAgente) + " Mes: " + str(mes) + " Ano: " + str(ano)
        else:
          ##print "Keys ja existem! Vamos fazer update do valor!"
          pass ##Nao executa nada
        cur.close()


        ##LISTA DE VALORES A SEREM INSERIDO - UPDATE
        valVenda = row[2].strip() ##Remove espacos

        if (not valVenda): ##Se Nao Ha' String (Vazia) - coloque 0.0
          valVenda = '0'

        valVenda = valVenda.replace(",",".") ##No Access inserimos com Decimal separado por PONTO
        lst_NomeCampo_Valor = []
        ##@@ (1)
        lst_NomeCampo_Valor.append("contrato_venda_mwh = " + str(valVenda))
        now = datetime.datetime.now()

        ##@@ (2)
        lst_NomeCampo_Valor.append("origem_contrato_venda = '" + "PUBLICO CCEE" + "'" )

        now = datetime.datetime.now()
        strNomeUsuario = getpass.getuser()
        ##@@ (3)
        ##lst_NomeCampo_Valor.append("QuemAtualizou = '" + strNomeUsuario + "'")

        ##@@ (4)
        lst_NomeCampo_Valor.append("DataAtualizacao_CVenda = '" + now.strftime("%d/%m/%Y %H:%M:%S") + "'")


        valores = ",".join(lst_NomeCampo_Valor)
        sql = sqlupd.replace("[ID_AGENTE]",str(idAgente))
        sql = sql.replace("[MES]", str(mes))
        sql = sql.replace("[ANO]", str(ano))
        sql = sql.replace("[VALORES]", valores)
        cur = conn.cursor()
        try:
            cur.execute(sql)
        except:
            print "Erro no UPDATE CVenda do XLSX!"
        conn.commit()
##------------------------------------------------------------------------------






################################################################################
## METODO @2014 - INSERE CONSUMO
################################################################################
def Insere_Medicao_Consumo_BD(conn,lstRowsTab,dtRef_RelatPublico,printDebug = False):
    ## CHECK - Keys
    sqlcheck = "SELECT idAgente, Mes, Ano FROM Agentes_Mensal_PublicoCCEE WHERE ((idAgente=[ID_AGENTE]) AND (Mes=[MES]) AND (mes=[MES]) AND (Ano=[ANO]))"
    ## INSERINDO Keys
    sqlins = "INSERT INTO Agentes_Mensal_PublicoCCEE(idAgente, Mes, Ano) VALUES ([ID_AGENTE], [MES], [ANO])"
    ## UPDATE - valor
    sqlupd = "UPDATE Agentes_Mensal_PublicoCCEE SET [VALORES] WHERE ((idAgente=[ID_AGENTE]) AND (Mes=[MES]) AND (Ano=[ANO]))"

    ##@@@@@
    ##@ (1.0) - DEBUG - IMPRIME O QUE SERA INSERIDO
##    if printDebug == True:
##      strFileNameLog =  "Saida/" + str(dtRef_RelatPublico.year) + "_" + str(dtRef_RelatPublico.month).zfill(2) + "_linhas_MedConsumo" + ".txt"
##      f = open(strFileNameLog, 'w')
##      for i in xrange(0,len(lstRowsTab)):
##        row = lstRowsTab[i]
##        val_PatLeve = row[2].replace("\xa0","").strip()
##        val_PatMedio = row[3].replace("\xa0","").strip()
##        val_PatPesado = row[4].replace("\xa0","").strip()
##
##        if (not val_PatLeve): ##Se Nao Ha String (Vazia) - coloque 0.0
##          val_PatLeve = '0'
##
##        if (not val_PatMedio):
##          val_PatMedio = '0'
##
##        if (not val_PatPesado):
##          val_PatPesado = '0'
##
##        val_PatLeve = val_PatLeve.replace(",",".")
##        val_PatMedio = val_PatMedio.replace(",",".")
##        val_PatPesado = val_PatPesado.replace(",",".")
##        valTotalCons = float(val_PatLeve) + float(val_PatMedio) + float(val_PatPesado)
##        strCol3_Valor = str(valTotalCons)
##        ##Para imprimir o CSV, substitui PONTO por VIRGULA
##        strCol3_Valor = strCol3_Valor.replace(".",",")
##        f.write(row[0] + ";" + row[1] + ";" + strCol3_Valor + "\n")
##      f.close()

    dia = dtRef_RelatPublico.day
    mes = dtRef_RelatPublico.month
    ano = dtRef_RelatPublico.year

    if printDebug == True:
      print "##########################################"
      print " INSERINDO CONSUMO (Html) (" + str(dtRef_RelatPublico.year) + "/" + str(dtRef_RelatPublico.month).zfill(2) + ")"
      print "##########################################"

    ##@@@@@
    ##@ (2.0) - LOOP - por TODAS as Linhas da TABELA
    intTotalReg = len(lstRowsTab)
    count = 0

    for row in lstRowsTab:
        count = count + 1
        if ( (count % 100) == 0):
          porcent = (float(count)/float(intTotalReg))
          print ("inserido %10.2f porc" % (float(porcent)*100.) )

        strSiglaAgente = ""
        strRazSocAgente = ""
        try:
          strSiglaAgente = str(row[0]).strip()
          strRazSocAgente = str(row[1]).strip() ##strip() Remove espacos do Inicio e Fim
        except:
            if printDebug == True:
              print "Linha em Branco - Pulando!"
            continue

        ##@@@@@
        ##@ (2.1) - Check Agente - pegando o ID do BD Local
        ## - se Nao Existir Insere na Tabela Agentes e Agentes_Sinonimos
        ## e MARCA que foi inserido do PublicoCCEE LIDO DO MES/ANO na origem_agente
        ## e MARCA que o CodCliq e' TEMPORARIO
##        if printDebug == 1:
##          print "LINHA MED CONS: " + str(count)
##          print "GET ID: " + strSiglaAgente

        idAgente = CheckAgente_FromTabSinonimos(conn, strSiglaAgente, dtRef_RelatPublico, True)

##        if printDebug == 1:
##          print "ID: " + strSiglaAgente + " = " + str(idAgente)

        sql = sqlcheck.replace("[ID_AGENTE]",str(idAgente))
        sql = sql.replace("[MES]", str(mes))
        sql = sql.replace("[ANO]", str(ano))
        cur = conn.cursor()

        cur.execute(sql)
        result = cur.fetchone()
        if (result == None):
          sql = sqlins.replace("[ID_AGENTE]",str(idAgente))
          sql = sql.replace("[MES]", str(mes))
          sql = sql.replace("[ANO]", str(ano))
          cur = conn.cursor()
          try:
              cur.execute(sql)
          except:
              print "ERRO na INSERCAO keys (MED CONSUMO) de Agentes_Mensal_PublicoCCEE ID= " + str(idAgente) + " Mes: " + str(mes) + " Ano: " + str(ano)
        else:
          pass
        cur.close()


        ##LISTA DE VALORES A SEREM INSERIDO - UPDATE
        val_PatLeve = row[2].replace("\xa0","").strip() ##Remove espacos e \xa0
        val_PatMedio = row[3].replace("\xa0","").strip() ##Remove espacos
        val_PatPesado = row[4].replace("\xa0","").strip() ##Remove espacos

        if (not val_PatLeve): #Se Nao Ha' String (Vazia) - coloque 0.0
          val_PatLeve = '0'

        if (not val_PatMedio):
          val_PatMedio = '0'

        if (not val_PatPesado):
          val_PatPesado = '0'

        val_PatLeve = val_PatLeve.replace(",",".")
        val_PatMedio = val_PatMedio.replace(",",".")
        val_PatPesado = val_PatPesado.replace(",",".")
        valTotalCons = float(val_PatLeve) + float(val_PatMedio) + float(val_PatPesado)

        if (not valTotalCons): ##Se Nao Ha' String (Vazia) - coloque 0.0
          valTotalCons = '0'

        lst_NomeCampo_Valor = []
        ##@@ (1)
        lst_NomeCampo_Valor.append("carga_consumo_mwh = " + str(valTotalCons))
        now = datetime.datetime.now()

        ##@@ (2)
        lst_NomeCampo_Valor.append("origem_consumo = '" + "PUBLICO CCEE" + str(dtRef_RelatPublico.year) + "/" + str(dtRef_RelatPublico.month).zfill(2)  + "'" )

        now = datetime.datetime.now()
        strNomeUsuario = getpass.getuser()
        ##@@ (3)
        ##lst_NomeCampo_Valor.append("QuemAtualizou = '" + strNomeUsuario + "'")
        ##@@ (4)
        lst_NomeCampo_Valor.append("DataAtualizacao_Consumo = '" + now.strftime("%d/%m/%Y %H:%M:%S") + "'")

        valores = ",".join(lst_NomeCampo_Valor)
        sql = sqlupd.replace("[ID_AGENTE]",str(idAgente))
        sql = sql.replace("[MES]", str(mes))
        sql = sql.replace("[ANO]", str(ano))
        sql = sql.replace("[VALORES]", valores)
        cur = conn.cursor()
        ##print sql
        try:
            cur.execute(sql)
        except:
            print "ERRO UPDATE (MED CONSUMO)  Tab Agentes_Mensal_PublicoCCEE ID= " + str(idAgente) + " Mes: " + str(mes) + " Ano: " + str(ano)
        conn.commit()
##------------------------------------------------------------------------------



################################################################################
## METODO @2014 - INSERE GARANTIA FISICA NO BD
##
## EXEMPLO: DADO UMA LISTA EXTRAIDA DO HTML COM idAgente,strNomeAg, Mes,Ano
## e valor - 1o Verifica se o SiglaAgente ja existe na Tabela de Sinonimos, caso
## exista apenas insere os dados na Tabela Agentes_Mensal_PublicoCCEE
## idAgente, Mes, Ano e o valor no campo: Venda
################################################################################
def Insere_GarantiaFisica_BD(conn,lstRowsTab,dtRef_RelatPublico,printDebug = False):
    ## CHECK - Keys
    sqlcheck = "SELECT idAgente, Mes, Ano FROM Agentes_Mensal_PublicoCCEE WHERE ((idAgente=[ID_AGENTE]) AND (Mes=[MES]) AND (mes=[MES]) AND (Ano=[ANO]))"
    ## INSERINDO Keys
    sqlins = "INSERT INTO Agentes_Mensal_PublicoCCEE(idAgente, Mes, Ano) VALUES ([ID_AGENTE], [MES], [ANO])"
    ## UPDATE - valor
    sqlupd = "UPDATE Agentes_Mensal_PublicoCCEE SET [VALORES] WHERE ((idAgente=[ID_AGENTE]) AND (Mes=[MES])  AND (Ano=[ANO]))"

    ##@@@@@
    ##@ (1.0) - DEBUG - IMPRIME O QUE SERA INSERIDO
##    if printDebug == True:
##      strFileNameLog = "Saida/" + str(dtRef_RelatPublico.year) + "_" + str(dtRef_RelatPublico.month).zfill(2) + "_linhas_GFisica" + ".txt"
##      f = open(strFileNameLog, 'w')
##      for i in xrange(0,len(lstRowsTab)):
##        row = lstRowsTab[i]
##        val_PatLeve = row[2].replace("\xa0","").strip()
##        val_PatMedio = row[3].replace("\xa0","").strip()
##        val_PatPesado = row[4].replace("\xa0","").strip()
##
##        if (not val_PatLeve): ##Se Nao Ha String (Vazia) - coloque 0.0
##          val_PatLeve = '0'
##
##        if (not val_PatMedio):
##          val_PatMedio = '0'
##
##        if (not val_PatPesado):
##          val_PatPesado = '0'
##
##        val_PatLeve = val_PatLeve.replace(",",".")
##        val_PatMedio = val_PatMedio.replace(",",".")
##        val_PatPesado = val_PatPesado.replace(",",".")
##        valTotalGF = float(val_PatLeve) + float(val_PatMedio) + float(val_PatPesado)
##        strCol3_Valor = str(valTotalGF)
##        ##Para imprimir o CSV, substitui PONTO por VIRGULA
##        strCol3_Valor = strCol3_Valor.replace(".",",")
##        f.write(row[0] + ";" + row[1] + ";" + strCol3_Valor + "\n")
##      f.close()

    dia = dtRef_RelatPublico.day
    mes = dtRef_RelatPublico.month
    ano = dtRef_RelatPublico.year

    if printDebug == True:
      print "##########################################"
      print " INSERINDO GARANTIA FISICA (Html) (" + str(dtRef_RelatPublico.year) + "/" + str(dtRef_RelatPublico.month).zfill(2) + ")"
      print "##########################################"

    ##@@@@@
    ##@ (2.0) - LOOP - por TODAS as Linhas da TABELA
    intTotalReg = len(lstRowsTab)
    count = 0

    for row in lstRowsTab:
        count = count + 1
        if ( (count % 100) == 0):
          porcent = (float(count)/float(intTotalReg))
          print ("inserido %10.2f porc" % (float(porcent)*100.) )

        strSiglaAgente = ""
        strRazSocAgente = ""
        try:
          strSiglaAgente = str(row[0]).strip()
          strRazSocAgente = str(row[1]).strip() ##strip() Remove espacos do Inicio e Fim
        except:
            if printDebug == True:
              print "Linha em Branco - Pulando!"
            continue

        ##@@@@@
        ##@ (2.1) - Check Agente - pegando o ID do BD Local
        ## - se Nao Existir Insere na Tabela Agentes e Agentes_Sinonimos
        ## e MARCA que foi inserido do PublicoCCEE LIDO DO MES/ANO na origem_agente
        ## e MARCA que o CodCliq e' TEMPORARIO
##        if printDebug == 1:
##          print "LINHA GF: " + str(count)
##          print "Pegando o ID do Agente: " + strSiglaAgente
        idAgente = CheckAgente_FromTabSinonimos(conn, strSiglaAgente, dtRef_RelatPublico, True)

##        if printDebug == 1:
##          print "ID do " + strSiglaAgente + " = " + str(idAgente)

        sql = sqlcheck.replace("[ID_AGENTE]",str(idAgente))
        sql = sql.replace("[MES]", str(mes))
        sql = sql.replace("[ANO]", str(ano))
        cur = conn.cursor()

        cur.execute(sql)
        result = cur.fetchone()
        if (result == None):
          sql = sqlins.replace("[ID_AGENTE]",str(idAgente))
          sql = sql.replace("[MES]", str(mes))
          sql = sql.replace("[ANO]", str(ano))
          cur = conn.cursor()
          try:
              cur.execute(sql)
          except:
              print "ERRO na keys de Agentes_Mensal_PublicoCCEE: " + str(idAgente) + " Mes: " + str(mes) + " Ano: " + str(ano)
        else:
          pass
        cur.close()


        ##LISTA DE VALORES A SEREM INSERIDO - UPDATE
        val_PatLeve = row[2].replace("\xa0","").strip() ##Remove espacos
        val_PatMedio = row[3].replace("\xa0","").strip() ##Remove espacos
        val_PatPesado = row[4].replace("\xa0","").strip() ##Remove espacos

        if (not val_PatLeve): #Se Nao Ha' String (Vazia) - coloque 0.0
          val_PatLeve = '0'

        if (not val_PatMedio):
          val_PatMedio = '0'

        if (not val_PatPesado):
          val_PatPesado = '0'

        val_PatLeve = val_PatLeve.replace(",",".")
        val_PatMedio = val_PatMedio.replace(",",".")
        val_PatPesado = val_PatPesado.replace(",",".")

        ##print val_PatLeve + ";" + val_PatMedio + ";" + val_PatPesado
        valTotalGF = float(val_PatLeve) + float(val_PatMedio) + float(val_PatPesado)

        if (not valTotalGF): ##Se Nao Ha' String (Vazia) - coloque 0.0
          valTotalGF = '0'


        lst_NomeCampo_Valor = []
        ##@@ (1)
        lst_NomeCampo_Valor.append("gf_mwh = " + str(valTotalGF))
        now = datetime.datetime.now()

        ##@@ (2)
        lst_NomeCampo_Valor.append("origem_gf = '" + "PUBLICO CCEE" + str(dtRef_RelatPublico.year) + "/" + str(dtRef_RelatPublico.month).zfill(2)  + "'" )

        now = datetime.datetime.now()
        strNomeUsuario = getpass.getuser()
        ##@@ (3)
        ##lst_NomeCampo_Valor.append("QuemAtualizou = '" + strNomeUsuario + "'")
        ##@@ (4)
        lst_NomeCampo_Valor.append("DataAtualizacao_GF = '" + now.strftime("%d/%m/%Y %H:%M:%S") + "'")

        valores = ",".join(lst_NomeCampo_Valor)
        sql = sqlupd.replace("[ID_AGENTE]",str(idAgente))
        sql = sql.replace("[MES]", str(mes))
        sql = sql.replace("[ANO]", str(ano))
        sql = sql.replace("[VALORES]", valores)
        cur = conn.cursor()
        try:
            cur.execute(sql)
        except:
            print "Erro no UPDATE do Registro!"
        conn.commit()
##------------------------------------------------------------------------------



################################################################################
## METODO I @2014 - INSERE CONTRATOS DE COMPRA NO BD - por HTML
################################################################################
def Insere_ContratosCompra_BD(conn,lstRowsTab,dtRef_RelatPublico,printDebug = False):
    ## CHECK - Keys
    sqlcheck = "SELECT idAgente, Mes, Ano FROM Agentes_Mensal_PublicoCCEE WHERE ((idAgente=[ID_AGENTE]) AND (Mes=[MES]) AND (mes=[MES]) AND (Ano=[ANO]))"
    ## INSERINDO Keys
    sqlins = "INSERT INTO Agentes_Mensal_PublicoCCEE(idAgente, Mes, Ano) VALUES ([ID_AGENTE], [MES], [ANO])"
    ## UPDATE - valor
    sqlupd = "UPDATE Agentes_Mensal_PublicoCCEE SET [VALORES] WHERE ((idAgente=[ID_AGENTE]) AND (Mes=[MES]) AND (Ano=[ANO]))"

    ##@@@@@
    ##@ (1.0) - DEBUG - IMPRIME O QUE SERA INSERIDO
    if printDebug == True:
      strFileNameLog = "Saida/" + str(dtRef_RelatPublico.year) + "_" + str(dtRef_RelatPublico.month).zfill(2) + "_linhas_CCOMPRA" + ".txt"
      f = open(strFileNameLog, 'w')
      for i in xrange(0,len(lstRowsTab)):
        row = lstRowsTab[i]
        strNomeAgemte = row[0].strip() ##Remove espacos
        strRazSocAgente = row[1].strip() ##Remove espacos
        strVal = row[2].replace("\xa0","").strip() ##Remove espacos
        if (not strVal): ##Se Nao Ha' String (Vazia) - coloque 0.0
          strVal = '0'
        strVal = strVal.replace(",",".")

        f.write(strNomeAgemte + ";" + strRazSocAgente + ";" + strVal + "\n")
      f.close()

    dia = dtRef_RelatPublico.day
    mes = dtRef_RelatPublico.month
    ano = dtRef_RelatPublico.year

    if printDebug == True:
      print "##########################################"
      print " INSERINDO CONTRATO DE COMPRA (Html) (" + str(dtRef_RelatPublico.year) + "/" + str(dtRef_RelatPublico.month).zfill(2) + ")"
      print "##########################################"

    ##@@@@@
    ##@ (2.0) - LOOP - por TODAS as Linhas da TABELA
    intTotalReg = len(lstRowsTab)
    count = 0

    for row in lstRowsTab:
        count = count + 1
        if printDebug == 1:
          print "Linha " + str(count) + " da lista de Contratos"

        if ( (count % 100) == 0):
          porcent = (float(count)/float(intTotalReg))
          print ("inserido %10.2f porc" % (float(porcent)*100.) )

        strSiglaAgente = ""
        strRazSocAgente = ""
        try:
          strSiglaAgente = str(row[0]).strip()
          strRazSocAgente = str(row[1]).strip() ##strip() Remove espacos do Inicio e Fim
        except:
            if printDebug == True:
              print "Linha em Branco - Pulando!"
            continue

        ##@@@@@
        ##@ (2.1) - Check Agente - pegando o ID do BD Local
        ## - se Nao Existir Insere na Tabela Agentes e Agentes_Sinonimos
        ## e MARCA que foi inserido do PublicoCCEE LIDO DO MES/ANO na origem_agente
        ## e MARCA que o CodCliq e' TEMPORARIO
##        if printDebug == 1:
##          print "Pegando ID para Agente: " + strSiglaAgente
        idAgente = CheckAgente_FromTabSinonimos(conn, strSiglaAgente, dtRef_RelatPublico, True)

##        if printDebug == 1:
##          print "ID do " + strSiglaAgente + " = " + str(idAgente)

        sql = sqlcheck.replace("[ID_AGENTE]",str(idAgente))
        sql = sql.replace("[MES]", str(mes))
        sql = sql.replace("[ANO]", str(ano))
        cur = conn.cursor()

        cur.execute(sql)
        result = cur.fetchone()
        if (result == None):
          sql = sqlins.replace("[ID_AGENTE]",str(idAgente))
          sql = sql.replace("[MES]", str(mes))
          sql = sql.replace("[ANO]", str(ano))
          cur = conn.cursor()
          try:
              if printDebug == 1:
                print "Tentando INSERIR: \n"
                print sql
              cur.execute(sql)
          except:
              print "ERRO na keys de INSERCAO - TABELA: Agentes_Mensal_PublicoCCEE IDAgente: " + str(idAgente) + " Mes: " + str(mes) + " Ano: " + str(ano)
        else:
          pass
        cur.close()

        ##LISTA DE VALORES A SEREM INSERIDO - UPDATE
        strVal = row[2].replace("\xa0","").strip() ##Remove espacos

        if (not strVal): ##Se Nao Ha' String (Vazia) - coloque 0.0
          strVal = '0'

        strVal = strVal.replace(",",".") ##No Access inserimos com Decimal separado por PONTO
        lst_NomeCampo_Valor = []
        ##@@ (1)
        lst_NomeCampo_Valor.append("contrato_compra_mwh = " + strVal)
        now = datetime.datetime.now()

        ##@@ (2)
        lst_NomeCampo_Valor.append("origem_contrato_compra = '" + "PUBLICO CCEE" + str(mes) + "/" + str(ano) + "'" )

        now = datetime.datetime.now()
        strNomeUsuario = getpass.getuser()
        ##@@ (3)
        ##lst_NomeCampo_Valor.append("QuemAtualizou = '" + strNomeUsuario + "'")
        ##@@ (4)
        lst_NomeCampo_Valor.append("DataAtualizacao_CCompra = '" + now.strftime("%d/%m/%Y %H:%M:%S") + "'")


        valores = ",".join(lst_NomeCampo_Valor)
        sql = sqlupd.replace("[ID_AGENTE]",str(idAgente))
        sql = sql.replace("[MES]", str(mes))
        sql = sql.replace("[ANO]", str(ano))
        sql = sql.replace("[VALORES]", valores)
        cur = conn.cursor()
        try:
            cur.execute(sql)
        except:
            print "Erro no UPDATE do Registro!"
        conn.commit()
##------------------------------------------------------------------------------









################################################################################
## METODO I @2014 - INSERE CONTRATOS DE VENDA NO BD - por HTML
##
################################################################################
def Insere_ContratosVenda_BD(conn,lstRowsTab,dtRef_RelatPublico,printDebug = False):
    ## CHECK - Keys
    sqlcheck = "SELECT idAgente, Mes, Ano FROM Agentes_Mensal_PublicoCCEE WHERE ((idAgente=[ID_AGENTE]) AND (Mes=[MES]) AND (mes=[MES]) AND (Ano=[ANO]))"
    ## INSERINDO Keys
    sqlins = "INSERT INTO Agentes_Mensal_PublicoCCEE(idAgente, Mes, Ano) VALUES ([ID_AGENTE], [MES], [ANO])"
    ## UPDATE - valor
    sqlupd = "UPDATE Agentes_Mensal_PublicoCCEE SET [VALORES] WHERE ((idAgente=[ID_AGENTE]) AND (Mes=[MES]) AND (mes=[MES]) AND (Ano=[ANO]))"

    ##@@@@@
    ##@ (1.0) - DEBUG - IMPRIME O QUE SERA INSERIDO
    if printDebug == True:
      strFileNameLog = "Saida/" +  str(dtRef_RelatPublico.year) + "_" + str(dtRef_RelatPublico.month).zfill(2) + "_linhas_CVenda" + ".txt"
      f = open(strFileNameLog, 'w')
      for i in xrange(0,len(lstRowsTab)):
        row = lstRowsTab[i]
        strNomeAgemte = row[0].strip() ##Remove espacos
        strRazSocAgente = row[1].strip() ##Remove espacos
        strValVenda = row[2].replace("\xa0","").strip() ##Remove espacos
        if (not strValVenda): ##Se Nao Ha' String (Vazia) - coloque 0.0
          strValVenda = '0'
        strValVenda = strValVenda.replace(",",".")

        f.write(strNomeAgemte + ";" + strRazSocAgente + ";" + strValVenda + "\n")
      f.close()

    dia = dtRef_RelatPublico.day
    mes = dtRef_RelatPublico.month
    ano = dtRef_RelatPublico.year

    if printDebug == True:
      print "##########################################"
      print " INSERINDO CONTRATO DE VENDA (Html) (" + str(dtRef_RelatPublico.year) + "/" + str(dtRef_RelatPublico.month).zfill(2) + ")"
      print "##########################################"

    ##@@@@@
    ##@ (2.0) - LOOP - por TODAS as Linhas da TABELA
    intTotalReg = len(lstRowsTab)
    count = 0

    for row in lstRowsTab:
        count = count + 1
        if ( (count % 100) == 0):
          porcent = (float(count)/float(intTotalReg))
          print ("inserido %10.2f porc" % (float(porcent)*100.) )

        strSiglaAgente = ""
        strRazSocAgente = ""
        try:
          strSiglaAgente = str(row[0]).strip()
          strRazSocAgente = str(row[1]).strip() ##strip() Remove espacos do Inicio e Fim
        except:
            if printDebug == True:
              print "Linha em Branco - Pulando!"
            continue

        ##@@@@@
        ##@ (2.1) - Check Agente - pegando o ID do BD Local
        ## - se Nao Existir Insere na Tabela Agentes e Agentes_Sinonimos
        ## e MARCA que foi inserido do PublicoCCEE LIDO DO MES/ANO na origem_agente
        ## e MARCA que o CodCliq e' TEMPORARIO
        idAgente = CheckAgente_FromTabSinonimos(conn, strSiglaAgente, dtRef_RelatPublico, True)


        sql = sqlcheck.replace("[ID_AGENTE]",str(idAgente))
        sql = sql.replace("[MES]", str(mes))
        sql = sql.replace("[ANO]", str(ano))
        cur = conn.cursor()

        cur.execute(sql)
        result = cur.fetchone()
        if (result == None):
          sql = sqlins.replace("[ID_AGENTE]",str(idAgente))
          sql = sql.replace("[MES]", str(mes))
          sql = sql.replace("[ANO]", str(ano))
          cur = conn.cursor()
          try:
              cur.execute(sql)
          except:
              print "ERRO na keys de Agentes_Mensal_PublicoCCEE: " + str(idAgente) + " Mes: " + str(mes) + " Ano: " + str(ano)
        else:
          pass
        cur.close()

        ##LISTA DE VALORES A SEREM INSERIDO - UPDATE
        valVenda = row[2].replace("\xa0","").strip() ##Remove espacos

        if (not valVenda): ##Se Nao Ha' String (Vazia) - coloque 0.0
          valVenda = '0'

        valVenda = valVenda.replace(",",".") ##No Access inserimos com Decimal separado por PONTO
        lst_NomeCampo_Valor = []
        ##@@ (1)
        lst_NomeCampo_Valor.append("contrato_venda_mwh = " + valVenda)
        now = datetime.datetime.now()

        ##@@ (2)
        lst_NomeCampo_Valor.append("origem_contrato_venda = '" + "PUBLICO CCEE" + "'" )

        now = datetime.datetime.now()
        strNomeUsuario = getpass.getuser()
        ##@@ (3)
        ##lst_NomeCampo_Valor.append("QuemAtualizou = '" + strNomeUsuario + "'")
        ##@@ (4)
        lst_NomeCampo_Valor.append("DataAtualizacao_CVenda = '" + now.strftime("%d/%m/%Y %H:%M:%S") + "'")


        valores = ",".join(lst_NomeCampo_Valor)
        sql = sqlupd.replace("[ID_AGENTE]",str(idAgente))
        sql = sql.replace("[MES]", str(mes))
        sql = sql.replace("[ANO]", str(ano))
        sql = sql.replace("[VALORES]", valores)
        cur = conn.cursor()
        try:
            cur.execute(sql)
        except:
            print "Erro no UPDATE do Registro!"
        conn.commit()
##------------------------------------------------------------------------------



################################################################################
### METODO: CHECK TAB SINONIMOS
## Verifica se AGENTE ja existe no PUBLICOCCEE
###@2014 - Agora temos que controlar o CodCliq e idAgente das Tabelas Agentes e
###        Agentes_Sinonimos.
###  - Pois os Nomes dos Agentes podem vir do Relatorio Mensal da CCEE com Todas
### as informacoes corretamente, ou do Relatorio Publico CCEE com info de Consumo
### de GF.. se POR ACASO o Agente informado no Publico CCEE nao existir no BD
### ele insere com COD_AGENTE MAIOR QUE 999999 - E indica que origem_nome = PUBLICO_CCEE
### Posteriormente, ao Atualizarmos a Lista de Agentes, faremos essa correcao
### colocando os Dados Corretos desse Agente - principalmente no codCliq dele.
## O argumento dtRefRelat_Publico -> passa a Informacao de Qual Mes/Ano
## estavamos lendo a informacao que tivemos que INSERIR o AGENTE NO BD -
## Assim conseguiremos rastrear melhor se Esse AGENTE e' muito antigo e SAIU DA LISTA
## MENSAL DE AGENTES DA CCEE
def CheckAgente_FromTabSinonimos(conn, strSiglaPerfilAgente ,dtRefRelat_Publico, printDebug = False):

        ##Insercao das Datas no Campo comentarios
        now = datetime.datetime.now()
        dtHoje = datetime.datetime(now.year, now.month, now.day)

        ##@[CHECK] - VEJA O PUBLICO CCEE FORNECE A SIGLA DO PERFIL DO AGENTE E NAO A SIGLA DO AGENTE
        sql_check_SINON = "SELECT idAgente FROM Agentes_Sinonimos WHERE (((SiglaPerfilAgente)='[SIGLA_PERFIL_AGENTE]'))"

        ##ID Local
        sql_get_MaxID = "SELECT MAX(idAgente) AS Maximo FROM Agentes"


        ##@[INSERT EM SINONIMOS]
        sql_insert_SINON =  "INSERT INTO Agentes_Sinonimos (idAgente, SiglaPerfilAgente, DataInsercao, origem_agente, QuemInseriu) \
         VALUES ([ID], '[SIGLA_PERFIL_AGENTE]', '[DATA_INS]','[ORIGEM]','[QUEM]')"

        ##@[INSERT EM AGENTES]
        sql_insert =  "INSERT INTO Agentes (idAgente, SiglaPerfilAgente, DataInsercao, origem_agente, QuemInseriu) \
         VALUES ([ID], '[SIGLA_PERFIL_AGENTE]', '[DATA_INS]','[ORIGEM]', '[QUEM]')"

        ##@@@@@
        ##@ (3.0) - MONTA O CHECK
        sql = sql_check_SINON.replace("[SIGLA_PERFIL_AGENTE]",strSiglaPerfilAgente)
        cur = conn.cursor()
        cur.execute(sql)

        idAgente = 0
        maxID = 0
        maxCodCliq = 0
        result = cur.fetchone()

        ##@@@@@
        ##@ CHECK - CONSULTA
        if (result == None): ##NOT FOUND
            strTextoErro =  now.strftime("%d/%m/%Y %H:%M:%S") + " "
            strTextoErro = strTextoErro + " Sigla_Perfil_Agente: '" + str(strSiglaPerfilAgente) + "' PUBLIC_CCEE " \
            + str(dtRefRelat_Publico.year) + "/" + str(dtRefRelat_Publico.month).zfill(2) + "\n"
            strFileNameLog = "Saida/ERROS_AGENTES_NAO_ENCONTRADOS_INSERIDOS_CODCLIQ_FALSO.log"
            if printDebug == True:
              print "@@@ AGENTE INSERIDO : " +  str(strSiglaPerfilAgente)
            write_LOG_ERROR(strFileNameLog, strTextoErro)

            ##@@@@@
            ##@ (3.1) - MAX ID
            cur.close()
            cur = conn.cursor()
            cur.execute(sql_get_MaxID)
            result = cur.fetchone() ##Pega o MAX(idAgente)
            if(result[0] != None):
              try:
                  maxID = int(result[0]) ##Max ID - some + 1 para Inserir um Novo
              except:
                  maxID = 0

            ##Se for inserir o codCliq abaixo comente o cur.close()
            cur.close()


            ##ID do NOVO AGENTE
            idAgente = maxID + 1

            ##Insere o SiglaAgente nas 2(duas) tabelas - TabSinonimo e Agentes
            if (maxID != 0):

                ##@[INSERT EM AGENTES]
                sql = sql_insert.replace("[SIGLA_PERFIL_AGENTE]",strSiglaPerfilAgente)
                sql = sql.replace("[ID]", str(idAgente))
                ##@@ (origem)
                sql = sql.replace("[ORIGEM]", "PUBLICO_CCEE DE " + str(dtRefRelat_Publico.month).zfill(2) + "/" + str(dtRefRelat_Publico.year) )
                ##@@ QUEM INSERIU
                sql = sql.replace("[QUEM]", getpass.getuser() )
                ##@@ QUANDO INSERIU
                sql = sql.replace("[DATA_INS]", now.strftime("%d/%m/%Y %H:%M:%S") )
                cur = conn.cursor()
                print sql
                cur.execute(sql)
                conn.commit()
                cur.close()

                ##@[INSERT EM SINONIMOS DE AGENTES]
                sql = sql_insert_SINON.replace("[SIGLA_PERFIL_AGENTE]",strSiglaPerfilAgente)
                sql = sql.replace("[ID]", str(idAgente))
                ##@@ (origem)
                sql = sql.replace("[ORIGEM]", "PUBLICO_CCEE DE " + str(dtRefRelat_Publico.month).zfill(2) + "/" + str(dtRefRelat_Publico.year) )
                ##@@ QUEM INSERIU
                sql = sql.replace("[QUEM]", getpass.getuser() )
                ##@@ QUANDO INSERIU
                sql = sql.replace("[DATA_INS]", now.strftime("%d/%m/%Y %H:%M:%S") )
                cur = conn.cursor()
                cur.execute(sql)
                conn.commit()
                cur.close()

            elif (maxID == 0):
                print "E' O PRIMEIRO AGENTE A SER INSERIDO NO BD - ERRO NAO TRATADO"
                print "INSIRA PELO MENOS UM AGENTE nas TABELAS Agentes e Agentes_Sinonimos MANUALMENTE \n"
                exit

        else:
            idAgente = result[0] ##Encontrou o ID Interno do BD

        return idAgente
##------------------------------------------------------------------------------




################################################################################
## METODO: imprime um LOG - append a linha
################################################################################
def write_LOG_ERROR(strFileNameLog,strTxt):
    try:
        f = open(strFileNameLog, "a")
        try:
          f.write(strTxt)
        finally:
            f.close()
    except IOError:
        pass
##------------------------------------------------------------------------------


def update_TabPubCCEE_WithZeros(conn,strIDAno):
  sql_select = "SELECT IdAgente, Ano, Mes, GF, Compra, Venda, Consumo FROM Agentes_Mensal_PublicoCCEE WHERE (((Ano)=[ANO_CONSULTA])) ORDER BY Ano, Mes"
  sql_select = sql_select.replace("[ANO_CONSULTA]", strIDAno)
  sql_update = "UPDATE Agentes_Mensal_PublicoCCEE SET [VALORES] WHERE ( (IdAgente=[IDAGENTE]) AND (Ano=[ANO]) AND (Mes=[MES]) )"

  cur = conn.cursor()
  cur.execute(sql_select)
  lst4Campos = ['GF','Compra','Venda','Consumo']
  strCheckFile =  "Tab_PublicoCCEE.txt"
  ##f = open(strCheckFile, 'w')
  ##Nao usar fetchall qdo forem muitos dados pois carrega tudo para a Memoria
  ##Pega todas as linhas da Consulta o BD
  lstRowsTabPubCCEE = cur.fetchall() #Retorna Varias Linhas
  cur.close()
  countRow = 1
  for row in lstRowsTabPubCCEE:
    idAg = row.IdAgente
    Ano = row.Ano
    Mes = row.Mes

    ##Montando
    lstValores = []
    ##A row = IdAgente, Mes, Ano, GF, Compra, Venda, Consumo
    ##Quero apenas fazer update de GF em diante
    nCampos = 4
    ##Leia da Coluna 3 (GF) em diante
    for i in xrange(0, nCampos):
      if (row[i+3]!=""): ## Diferente de Zero
        lstValores.append(lst4Campos[i] + "= " + str(row[i+3]).replace(",","."))
      else:
        lstValores.append(lst4Campos[i] + "= 0.0")

    valores = ",".join(lstValores)
    ##Mesmo assim vieram os valores como None - vou colocar zero
    valores = valores.replace("None","0.0")

    sql = sql_update.replace("[IDAGENTE]",str(row.IdAgente))
    sql = sql.replace("[MES]", str(row.Mes))
    sql = sql.replace("[ANO]", str(row.Ano))
    sql = sql.replace("[VALORES]", valores)

    cur = conn.cursor()
    try:
        cur.execute(sql)
    except:
        print "Erro no UPDATE do Registro!"
    conn.commit()


  ##Final do Looping
  ##f.close()


##################################################################
## METODO - ESCREVE EXCEL 2007 - OPENPYXL
## Escreve Planilha com os Dados do BDO
##
##################################################################
def writeXLSX_BDO(conn,dtRefBDO,dest_filename,strDirAndFileName):

  ##DATA DE REFERENCIA DOS DADOS A SEREM INSERIDOS NO BD
  ##Inicializa a Data
  dtRefDados = datetime.datetime.now()
  lstUsinasBDO = []

  sql_lstUsinasBDO = "SELECT TabUsinasTermBDO.idUsinaTermica, TabUsinasTermBDO.codBTD, TabUsinasTermBDO.NomeUsinaTermica FROM TabUsinasTermBDO ORDER BY TabUsinasTermBDO.NomeUsinaTermica"

  f = open("Saida/Lst_Usinas.txt", "w")
  cur = conn.cursor()
  cur.execute(sql_lstUsinasBDO)
  #result = cur.fetchone()  # Retorna um Resultado
  #if (result != None):

  ##OBS:>
  ##Nao usar fetchall qdo forem muitos dados pois carrega tudo para a Memoria
  wb = Workbook()
  ##dest_filename = r'empty_book.xlsx'
  ws = wb.worksheets[0]
  ##Nomeia a Aba
  ws.title = "BD"

  ##Pega todas as linhas da Consulta o BD
  lstRows = cur.fetchall() ##>> SE FOREM MUITOS DADOS - CUIDADO - COLOCA TUDO NA MEM
  idRow = 0
  idCol = 0
  for row in lstRows:
      ##Converte para UTF-8 pois tem nome com acento
      strNomeUsina = row.NomeUsinaTermica.encode('utf-8')
      f.write(str(row.idUsinaTermica) + ";" + str(row.codBTD) + ";" + strNomeUsina + "\n")
      print row.idUsinaTermica, row.codBTD, strNomeUsina

      ws.cell(row = idRow, column = 1).value = str(row.idUsinaTermica)
      ws.cell(row = idRow, column = 2).value = str(row.codBTD)
      ws.cell(row = idRow, column = 3).value = strNomeUsina
      idRow = idRow + 1

  f.close()

##  for col_idx in xrange(1, 40):
##      col = get_column_letter(col_idx)
##      for row in xrange(1, 600):
##          ws.cell('%s%s'%(col, row)).value = '%s%s' % (col, row)

  ws = wb.create_sheet()
##  ws.title = 'Pi'
##  ws.cell('F5').value = 3.14
  wb.save(filename = dest_filename)

  print "### TERMINADO A ESCRITA DO EXCEL ###"















##################################################################
## Leitura dos Contratos em XLS dos Meses anteriores a MAR/2013
## A CCEE disponibiliza em XLS os valores
##
## Obs: vc pode simplificar a carga do excel colocando
##from openpyxl.reader.excel import Workbook
##from openpyxl.reader.excel import load_workbook
##################################################################
import string

def readConteudoXLS_ContratosCompra(idRowIni,idColIni,nColsTab,strFileName,strDirAndFileName):

  ##DATA DE REFERENCIA DOS DADOS A SEREM INSERIDOS NO BD
  ##Inicializa a Data
  dtRefDados = datetime.datetime.now()

  lstRowsTableCompra = []
  try:
    ##Planilhas muito Grandes a leitura rapida devemos usar iteradores
    ## Read Only e nao pode usar ws.cell(row = 10-1, column = 2-1).address
    ##wkb = openpyxl.reader.excel.load_workbook(strDirAndFileName, use_iterators = True)

    ##Carregando sem iteradores para podermos pegar o Endereco da Ultima linha
    ##da planilha
    wkb = openpyxl.reader.excel.load_workbook(strDirAndFileName, use_iterators = False)
    print strFileName + " foi carregado sem iteradores!"
    strRangeTab2Compra = ''
    try:
      ## Pegando todos os Nomes de Abas
      #wkshtList = wkb.get_sheet_names()
      #for wsName in wkshtList:
      #  print wsName

      ##Sem iteradores use_iterators = False
      wkshtList = wkb.get_sheet_names()
      ##print "Acessando a aba: " + wkshtList[0]
      ws = wkb.get_sheet_by_name(wkshtList[0])
      ##print "A ultima linha e ultima coluna da planilha e':"
      #print ws.get_highest_row()
      #print ws.get_highest_column()

      ##OBJETIVO: DADO O idRowIni e idColIni = (10,6) = F10
      ##MONTAR o TAMANHO da TABELA DE VENDAS
      ## idRow = 10 (linha 10, 10-1 0idx) e Coluna B = 2 (2-1 0idx)
      strAddressIni = ws.cell(row = int(idRowIni), column = int(idColIni)).address
      ##OBS: se a tabela vai ate a coluna D, devemos passar ate a coluna E para funcionar
      ##porntanto vamos usar nColsTab+1
      strAddressEnd = ws.cell(row = ws.get_highest_row(), column = int(idColIni)-1 + int(nColsTab)).address
      ## RANGE DA TABELA DE VENDAS
      ## strRangeTab2Compra = "F10:H2915"
      strRangeTab2Compra = strAddressIni + ":" + strAddressEnd

     ##Relatório de Informações ao Público - janeiro de 2013
      txtData = ws.cell("B2").value
      ##Monta um dicionario que dado o nome do mes retorna o id
      dic_NomeMes_ID = {}
      for i in xrange(1,13):
        strNomeMes = getNomeMes(i)
        dic_NomeMes_ID[strNomeMes] = i
        ##print strNomeMes + " = " + str(i)


      txtData = txtData.split('-')
      txtData = txtData[-1] ##remove espacos inicio e fim
      txtData = txtData.replace("-","").strip()
      txtData = txtData.encode('latin-1') ##Mes de Marco (cedilha)
      ##print txtData

      #- SEPARANDO A DATA
      #Divide no / Armazenando o Mes e Ano (ex: janeiro de 2013)
      strDataRef = str(txtData)
      (strMesExtenso,strAno) = strDataRef.split(' de ')
      strMesExtenso = strMesExtenso.strip()
      strAno = strAno.strip()
      idMesRef = dic_NomeMes_ID[strMesExtenso]
      dtRefDados = datetime.datetime(int(strAno),int(idMesRef),1)

    finally:
      pass

    ## - ABRINDO APENAS READONLY - COM ITERADORES - PARA ACELERAR LEITURA
    wkbWithIter = openpyxl.reader.excel.load_workbook(strDirAndFileName, use_iterators = True, guess_types=False, data_only=True)
    print strFileName + " foi carregado COM iteradores!"
    print "Lendo o Range dos Contratos de Compra: " + strRangeTab2Compra
    ##strRangeTab1Venda = '' ja foi preenchido acima
    try:
      wkshtList = wkbWithIter.get_sheet_names()
      ##print "Acessando a aba: " + wkshtList[0]

      ## DEFININDO A PLANILHA QUE FAREMOS A LEITURA
      ws = wkbWithIter.get_sheet_by_name(wkshtList[0])
      ##fout = open("Tab_Contrato_Compra_XLSX.txt", 'w')
      ## Looping por todas as linhas da TABELA -
      ## para acessarmos as colunas (3 no total) fazemos row[0],row[1] e row[2]
      for row in ws.iter_rows(range_string=str(strRangeTab2Compra)):

        ##if row[0].internal_value == openpyxl.cell.Cell.TYPE_STRING:
        ##  print row[0].encoding
        ##OBS: 1o Converte em String - depois faz o encode
        strCol1 = row[0].value ##Ja sabemos ser uma string, nao tente converter str()
        strCol2 = row[1].value ##pois dara problemas
        ##3a Coluna e' um Numero
        strCol3 = str(row[2].value)

        ## Como a Tabela de Compras tem mais linhas que a de Venda
        ## devemos testar se a 1a Coluna e' nula, com isso saimos
        ## do loop
        if (not strCol1):
          break ##Termina o Loop

        ##1o armazenamos a String em strCol1, depois
        ## interpretamos para o conjunto de caracteres
        ##Se tentamos fazer os 2 juntos apresenta problema
        strCol1 = strCol1.encode('latin-1').strip()
        strCol2 = strCol2.encode('latin-1').strip()
        strCol3 = strCol3.encode('latin-1').strip()

        ##Tem uns nomes que eles colocam .......... no final - Removendo
        strCol1 = strCol1.replace("..","")
        strCol2 = strCol2.replace("..","")

        ##.strip() ##Remove espacos do Inicio e Fim
        ##lstRowsTableCompra.append([row[0].internal_value,row[1].internal_value,row[2].internal_value])
        lstRowsTableCompra.append([strCol1,strCol2,strCol3])
        ##print "[" + strCol1 + ";" + strCol2+ ";" + str(strCol3) + "]"
        ##fout.write("[" + strCol1 + ";" + strCol2 + ";" + str(strCol3) + "]\n")

      ##fout.close()


    finally:
      pass

  ##Try Geral
  except IOError:
      print IOError.message


  return lstRowsTableCompra,dtRefDados


##################################################################
## Leitura dos Contratos em XLS dos Meses anteriores a MAR/2013
## A CCEE disponibiliza em XLS os valores
##
## Obs: vc pode simplificar a carga do excel colocando
##from openpyxl.reader.excel import Workbook
##from openpyxl.reader.excel import load_workbook
##
## E entao usar: wkb = load_workbook(strDirAndFileNameXLS)
## Dica:
      ##Pegando a Sheet que esta em 1o na Lista de Nomes
##      ##ws = workbook.get_sheet_by_name('Plan1')
##      ws = workbook.get_sheet_by_name(wkshtList[0])
##      ##ws agora e' um iterador
##      ##http://pythonhosted.org/openpyxl/optimized.html
##      ##cell, range, rows, columns methods and properties are disabled
##      for row in ws.iter_rows(): # it brings a new method: iter_rows()
##        for cell in row:
##          print cell.internal_value
      ##Lendo Aba Especifica e Valor Especifico
      ##sheet_ranges = wb.get_sheet_by_name(name = 'range names')
      ##print sheet_ranges.cell('D18').value # D18

      ##When a worksheet is created in memory, it contains no cells.
      ##They are created when first accessed. This way we don’t create
      ##objects that would never be accessed, thus reducing the memory footprint.
      ##http://pythonhosted.org/openpyxl/tutorial.html
##for i in xrange(0,100):
##...             for j in xrange(0,100):
##...                     ws.cell(row = i, column = j)
##will create 100x100 cells in memory, for nothing!!!!!!!!!!!!!!!!

##http://stackoverflow.com/questions/11162355/move-to-adjacent-cells-using-openpyxl

##Forum:https://groups.google.com/forum/#!forum/openpyxl-users
##http://pythonhosted.org/openpyxl/api.html?highlight=highest?highlight=highest#openpyxl.worksheet.Worksheet.get_highest_row

##https://bitbucket.org/wtking/openpyxl
##################################################################
import string

def readConteudoXLS_ContratosVenda(idRowIni,idColIni,nColsTab,strFileName,strDirAndFileName, printDebug = False):

  dtRefDados = datetime.datetime.now()
  lstRowsTableVenda = []
  try:
    ##Planilhas muito Grandes a leitura rapida devemos usar iteradores
    ## Read Only e nao pode usar ws.cell(row = 10-1, column = 2-1).address
    ##wkb = openpyxl.reader.excel.load_workbook(strDirAndFileName, use_iterators = True)

    ##Carregando sem iteradores para podermos pegar o Endereco da Ultima linha
    ##da planilha
    wkb = openpyxl.reader.excel.load_workbook(strDirAndFileName, use_iterators = False)
    print strFileName + " foi carregado sem iteradores!"
    strRangeTab1Venda = ''
    try:
      ## Pegando todos os Nomes de Abas
      ##wkshtList = wkb.get_sheet_names()
      ##for wsName in wkshtList:
      ##  print wsName
      wkshtList = wkb.get_sheet_names()
      ws = wkb.get_sheet_by_name(wkshtList[0])
      ##print "A ultima linha e ultima coluna da planilha e':"
      if printDebug == True:
        print "Acessando a aba: " + wkshtList[0]
        print "A ultima linha e Ultima coluna da planilha: "
        print ws.get_highest_row()
        print ws.get_highest_column()

      ##OBJETIVO: DADO O idRowIni e idColIni = (10,2) = B10
      strAddressIni = ws.cell(row = int(idRowIni), column = int(idColIni)).address
      ##OBS: se a tabela vai ate a coluna D, devemos passar ate a coluna E para funcionar
      ##porntanto vamos usar nColsTab+1
      strAddressEnd = ws.cell(row = ws.get_highest_row(), column = int(idColIni)-1 + int(nColsTab)).address

      ## RANGE DA TABELA DE VENDAS
      ## strRangeTab1Venda = "B10:E956"
      strRangeTab1Venda = strAddressIni + ":" + strAddressEnd

      ##Relatório de Informações ao Público - janeiro de 2013
      txtData = ws.cell("B2").value
      ##Monta um dicionario que dado o nome do mes retorna o id
      dic_NomeMes_ID = {}
      for i in xrange(1,13):
        strNomeMes = getNomeMes(i)
        dic_NomeMes_ID[strNomeMes] = i
        ##print strNomeMes + " = " + str(i)

      txtData = txtData.split('-')
      txtData = txtData[-1] ##remove espacos inicio e fim
      txtData = txtData.replace("-","").strip()
      txtData = txtData.encode('latin-1') ##Mes de Marco (cedilha)
      ##print txtData

      #- SEPARANDO A DATA
      #Divide no / Armazenando o Mes e Ano (ex: janeiro de 2013)
      strDataRef = str(txtData)
      (strMesExtenso,strAno) = strDataRef.split(' de ')
      strMesExtenso = strMesExtenso.strip()
      strAno = strAno.strip()
      idMesRef = dic_NomeMes_ID[strMesExtenso]
      dtRefDados = datetime.datetime(int(strAno),int(idMesRef),1)

    finally:
      pass


    ## - ABRINDO APENAS READONLY - COM ITERADORES - PARA ACELERAR LEITURA
    wkbWithIter = openpyxl.reader.excel.load_workbook(strDirAndFileName, use_iterators = True, guess_types=False, data_only=True)
    print strFileName + " COM iteradores - Contratos de VENDA!" + strRangeTab1Venda
    ##strRangeTab1Venda = '' ja foi preenchido acima
    try:
      wkshtList = wkbWithIter.get_sheet_names()
      print "Acessando a aba: " + wkshtList[0]
      ## DEFININDO A PLANILHA QUE FAREMOS A LEITURA
      ws = wkbWithIter.get_sheet_by_name(wkshtList[0])

      ##strRangeTab1Venda = "B10:D2914"
      for row in ws.iter_rows(range_string=str(strRangeTab1Venda)):

        ##if row[0].internal_value -->> internal value da O INDICE!

        ##OBS: 1o Converte em String - depois faz o encode
        strCol1 = row[0].value ##Ja sabemos ser uma string, nao tente converter str()
        strCol2 = row[1].value ##pois dara problemas
        ##3a Coluna e' um Numero
        strCol3 = str(row[2].value)

        ## Como a Tabela de Compras tem mais linhas que a de Venda
        ## devemos testar se a 1a Coluna e' nula, com isso saimos
        ## do loop
        if (not strCol1):
          break ##Termina o Loop

        ##1o armazenamos a String em strCol1, depois
        ## interpretamos para o conjunto de caracteres
        ##Se tentamos fazer os 2 juntos apresenta problema
        strCol1 = strCol1.encode('latin-1').strip()
        strCol2 = strCol2.encode('latin-1').strip()
        strCol3 = strCol3.encode('latin-1').strip()
        strCol1 = strCol1.replace("..","")
        strCol2 = strCol2.replace("..","")
        lstRowsTableVenda.append([strCol1,strCol2,strCol3])


    finally:
      pass
  ##Try Geral
  except IOError:
      print IOError.message


  if printDebug == True:
    strFileNameLog =  "Saida/" + str(dtRef_RelatPublico.year) + "_" + str(dtRef_RelatPublico.month).zfill(2) + "_CVenda_XLSX" + ".txt"
    f = open(strFileNameLog, 'w')
    for lstCellsRow in ws.range(strRangeTab1Venda):
      ##nCols = len(lstCellsRow)
      for j in xrange(0,len(lstCellsRow)):
        ## Varre as 3 colunas
        f.write( lstCellsRow[j].value )
        f.write(";")
      f.write("\n")
    f.close()


  return lstRowsTableVenda,dtRefDados



###############################################################################
#### METODO I - MEDIÇÃO - GERACAO - TABELA 1
## @autor PALIN
## Description: Extrai apenas a TABELA 1 - SOMA OS VALORES DOS PATAMARES
## strExpReg_TABConsumo = 'Perfil SiglaAgente(.*?)Tabela 2'
###############################################################################
def getLstRows_TabMedicao_GERACAO_FINAL_HtmlLocal(strConteudoHTML,printDebug = False):

  #Procura por uma data do tipo: 2013/03
  expRegData = re.compile(r'(([0-9]{4})/([0-9]{2}))')
  strDataRef = ''
  match = re.search(expRegData, strConteudoHTML)
  if match:
      strDataRef = match.group()
  #Divide no / Armazenando o Mes e Ano
  strDataRef = str(strDataRef)
  (strAno,strMes) = strDataRef.split('/')

  ##DATA DE REFERENCIA DOS DADOS A SEREM INSERIDOS NO BD
  dtRefDados = datetime.datetime(int(strAno),int(strMes),1)

  ##1 PASSO
  ## A PAGINA DA CCEE EXISTE VARIAS TABELAS DENTRO DE TABELAS
  ## VAMOS PEGAR APENAS O HTML REFERENTE A TABELA QUE QUEREMOS USANDO
  ## EXTRACAO POR EXPRESSAO REGULAR:
  strExpReg_TAB = 'Tabela 1(.*?)Tabela 2'
  strHtmlTab = ''
  for strBlocoFound in re.findall(strExpReg_TAB, strConteudoHTML, re.S):
    strHtmlTab = strBlocoFound

  ##(2) PASSO, dado que a expressao regular extraiu a partir da primeira
  idx0InicioTab2 = strHtmlTab.find("Empresa")
  strTabFinal1 = strHtmlTab[idx0InicioTab2:]

  strInicioHtmlTab = "<table><tbody><tr><td> Empresa"

  ##(3) Tabela reconstruida. Falta eliminar uma parte final apos encontrarmos
  ## o 1o </tbody></table>
  strTabFinal1 = strInicioHtmlTab + str(strTabFinal1)

  ##Encontrando o Fim
  idx0FinalTab = strTabFinal1.find("</tbody></table>")

  ##PRONTO - TENHO UM HTML APENAS COM A TABELA - VENDA
  ##Pegando + 17 caracteres para englobar a string </tbody></table>
  strTabFinal = strTabFinal1[:idx0FinalTab+17]

  ##DEBUG - Escreve um TXT com o HTML extraido
  if printDebug == True:
    try:
      f = open("Tabela_Geracao_Relat_Medicao.txt", "w")
      try:
        f.writelines(strTabFinal) # Escreve uma sequencia de Strings para o Arquivo
      finally:
        f.close()
    except IOError:
      pass

  ##(4) - Avaliando a Tabela (html) com BeautifulSoup
  soup = BeautifulSoup(strTabFinal)

  ##(5) - BeautifulSoup encontra no Html quantas tabelas existem
  ##desejamos que ele encontre apenas 1 tabela
  lstTables = soup.find_all('table')
  if printDebug == True:
    print "Existem " + str(len(lstTables)) + " tabelas na pagina"

  ##Apos avaliarmos linha a linha, coluna a coluna
  lstRowsTab = []

  ## INICIO DA LEITURA - CADA LINHA LE TODAS AS COLUNAS
  rows = lstTables[0].findAll('tr')
  for tr in rows:
    cols = tr.findAll('td')
    lstCols = []
    for td in cols:
        strText = td.find(text=True)
        strText = strText.encode('latin-1').strip()
        ##Dado que o numero esta escrito: 1.454,56, apenas
        ##removemos o .(ponto)
        strText = strText.replace(".","")
        strText = strText.replace("\r","")
        strText = strText.replace("\n","")
        strText = strText.replace("\t","")
        ##Colocamos na lista o caracter Windows ANSI
        ##lstCols.append(strText.encode('cp1252')) ##Trata caracteres acentuados
        lstCols.append(strText)
    lstRowsTab.append(lstCols)

  ## DEBUG - IMPRIME AS LINHAS EXTRAIDAS DA TABELA
  if printDebug == True:
    try:
      strFileName = "Saida/" + str(dtRefDados.year) + "_" + str(dtRefDados.month).zfill(2) + "Grab_HTML_Geracao.txt"
      f = open(strFileName, "w")
      try:
        for i in xrange(0,len(lstRowsTab)):
          row = lstRowsTab[i]
          for i in xrange(0,len(row)):
            f.write(row[i] + ";")
          f.write("\n")
      finally:
        f.close()
    except IOError:
      pass

  ##Retira o Cabecalho da Lista
  return lstRowsTab[1:],dtRefDados
############################# FINAL DO getLstRows_TabGarantiaFisica_HtmlLocal






###############################################################################
#### METODO I - MEDIÇÃO - CONSUMO - Tabela 3 - CONSUMO TOTAL
## @autor PALIN
## Description: Extrai apenas a TABELA 3 - Ajustada - SOMA AS GF L,M e P
## strExpReg_TABConsumo = 'Perfil SiglaAgente(.*?)Tabela 2'
###############################################################################
def getLstRows_TabMedicaoConsumo_HtmlLocal(strConteudoHTML,printDebug = False):

  #Procura por uma data do tipo: 2013/03
  expRegData = re.compile(r'(([0-9]{4})/([0-9]{2}))')
  strDataRef = ''
  match = re.search(expRegData, strConteudoHTML)
  if match:
      strDataRef = match.group()
  #Divide no / Armazenando o Mes e Ano
  strDataRef = str(strDataRef)
  (strAno,strMes) = strDataRef.split('/')

  ##DATA DE REFERENCIA DOS DADOS A SEREM INSERIDOS NO BD
  dtRefDados = datetime.datetime(int(strAno),int(strMes),1)

  ##1 PASSO
  ## A PAGINA DA CCEE EXISTE VARIAS TABELAS DENTRO DE TABELAS
  ## VAMOS PEGAR APENAS O HTML REFERENTE A TABELA QUE QUEREMOS USANDO
  ## EXTRACAO POR EXPRESSAO REGULAR:
  strExpReg_TAB = 'Tabela 3(.*?)Tabela 4'
  strHtmlTab = ''
  for strBlocoFound in re.findall(strExpReg_TAB, strConteudoHTML, re.S):
    strHtmlTab = strBlocoFound

  ##(2) PASSO, dado que a expressao regular extraiu a partir da primeira
  idx0InicioTab2 = strHtmlTab.find("Empresa")
  strTabFinal1 = strHtmlTab[idx0InicioTab2:]

  strInicioHtmlTab = "<table><tbody><tr><td> Empresa"

  ##(3) Tabela reconstruida. Falta eliminar uma parte final apos encontrarmos
  ## o 1o </tbody></table>
  strTabFinal1 = strInicioHtmlTab + str(strTabFinal1)

  ##Encontrando o Fim
  idx0FinalTab = strTabFinal1.find("</tbody></table>")

  ##PRONTO - TENHO UM HTML APENAS COM A TABELA - VENDA
  ##Pegando + 17 caracteres para englobar a string </tbody></table>
  strTabFinal = strTabFinal1[:idx0FinalTab+17]

  if printDebug == True:
    strFileName = "Saida/" + str(dtRefDados.year) + "_" + str(dtRefDados.month).zfill(2) + "_Medicao_Consumo.htm"
    try:
      f = open(strFileName, "w")
      try:
        f.write(strTabFinal)
      finally:
        f.close()
    except IOError:
      pass

  ##(4) - Avaliando a Tabela (html) com BeautifulSoup
  soup = BeautifulSoup(strTabFinal)

  ##(5) - BeautifulSoup encontra no Html quantas tabelas existem
  ##desejamos que ele encontre apenas 1 tabela
  lstTables = soup.find_all('table')
  if printDebug == True:
    print "Existem " + str(len(lstTables)) + " tabelas na pagina"

  ##Apos avaliarmos linha a linha, coluna a coluna
  lstRowsTab = []

  ## INICIO DA LEITURA - CADA LINHA LE TODAS AS COLUNAS
  rows = lstTables[0].findAll('tr')
  for tr in rows:
    cols = tr.findAll('td')
    lstCols = []
    for td in cols:
        strText = td.find(text=True)
        strText = strText.encode('latin-1').strip()
        ##Dado que o numero esta escrito: 1.454,56, apenas
        ##removemos o .(ponto)
        strText = strText.replace(".","")
        strText = strText.replace("\r","")
        strText = strText.replace("\n","")
        strText = strText.replace("\t","")
        ##Colocamos na lista o caracter Windows ANSI
        ##lstCols.append(strText.encode('cp1252')) ##Trata caracteres acentuados
        lstCols.append(strText)
    lstRowsTab.append(lstCols)

  ## DEBUG - IMPRIME AS LINHAS EXTRAIDAS DA TABELA
  if printDebug == True:
    try:
      strFileName = "Saida/" + str(dtRefDados.year) + "_" + str(dtRefDados.month).zfill(2) + "_lista_Medicao_Consumo.txt"
      f = open(strFileName, "w")
      try:
        for i in xrange(0,len(lstRowsTab)):
          row = lstRowsTab[i]
          for i in xrange(0,len(row)):
            f.write(row[i] + ";")
          f.write("\n")
      finally:
        f.close()
    except IOError:
      pass

  ##Retira o Cabecalho da Lista
  return lstRowsTab[1:],dtRefDados
############################# FINAL DO getLstRows_TabGarantiaFisica_HtmlLocal




###############################################################################
#### METODO INTERPRETA HTML - LEITURA - GARANTIA FISICA
## @autor PALIN
## Description: Extrai apenas a TABELA 2 - Ajustada - SOMA AS GF L,M e P
## strExpReg_TABConsumo = 'Perfil SiglaAgente(.*?)Tabela 2'
###############################################################################
def getLstRows_TabGarantiaFisica_HtmlLocal(strConteudoHTML,printDebug = False):

  #Procura por uma data do tipo: 2013/03
  expRegData = re.compile(r'(([0-9]{4})/([0-9]{2}))')
  strDataRef = ''
  match = re.search(expRegData, strConteudoHTML)
  if match:
      strDataRef = match.group()

  #- SEPARANDO A DATA
  #Divide no / Armazenando o Mes e Ano
  strDataRef = str(strDataRef)
  (strAno,strMes) = strDataRef.split('/')

  ##DATA DE REFERENCIA DOS DADOS A SEREM INSERIDOS NO BD
  dtRefDados = datetime.datetime(int(strAno),int(strMes),1)
  #print " Data do Dados: " + dtRefDados.strftime('%m/%Y')

  ##1 PASSO
  ## A PAGINA DA CCEE EXISTE VARIAS TABELAS DENTRO DE TABELAS
  ## VAMOS PEGAR APENAS O HTML REFERENTE A TABELA QUE QUEREMOS USANDO
  ## EXTRACAO POR EXPRESSAO REGULAR:
  ##strExpReg_TAB = 'Tabela 2(.*?)Linhas 1'

  ##Correcao 2014/Agosto - a palavra Linhas 1 nao existe em alguns relatorios
  strExpReg_TAB = 'Tabela 2(.*?)gina 1'
  strHtmlTab = ''
  for strBlocoFound in re.findall(strExpReg_TAB, strConteudoHTML, re.S):
    strHtmlTab = strBlocoFound


  ##(2) PASSO, dado que a expressao regular extraiu a partir da primeira
  idx0InicioTab2 = strHtmlTab.find("Perfil Agente")
  strTabFinal1 = strHtmlTab[idx0InicioTab2:]

  strInicioHtmlTab = "<table><tbody><tr><td> Perfil Agente"

  ##(3) Tabela reconstruida. Falta eliminar uma parte final apos encontrarmos
  ## o 1o </tbody></table>
  strTabFinal1 = strInicioHtmlTab + str(strTabFinal1)

  ##Encontrando o Fim
  idx0FinalTab = strTabFinal1.find("</tbody></table>")

  ##PRONTO - TENHO UM HTML APENAS COM A TABELA - VENDA
  ##Pegando + 17 caracteres para englobar a string </tbody></table>
  strTabFinal = strTabFinal1[:idx0FinalTab+17]

  if printDebug == True:
    strFileName = "Saida/" + str(dtRefDados.year) + "_" + str(dtRefDados.month).zfill(2) + "_Garantia_Fisica.htm"
    try:
      f = open(strFileName, "w")
      try:
        f.write(strTabFinal)
      finally:
        f.close()
    except IOError:
      pass

  ##(4) - Avaliando a Tabela (html) com BeautifulSoup
  soup = BeautifulSoup(strTabFinal)

  ##(5) - BeautifulSoup encontra no Html quantas tabelas existem
  ##desejamos que ele encontre apenas 1 tabela
  lstTables = soup.find_all('table')
  print "Existem " + str(len(lstTables)) + " tabelas na pagina"

  ##Apos avaliarmos linha a linha, coluna a coluna
  lstRowsTab = []

  ## INICIO DA LEITURA - CADA LINHA LE TODAS AS COLUNAS
  rows = lstTables[0].findAll('tr')
  for tr in rows:
    cols = tr.findAll('td')
    lstCols = [] #Limpa a lista de Colunas para Proxima Linha
    for td in cols:
        strText = td.find(text=True)
        strText = strText.encode('latin-1').strip()
        strText = strText.replace(".","")
        strText = strText.replace("\r","")
        strText = strText.replace("\n","")
        strText = strText.replace("\t","")
        lstCols.append(strText)
    lstRowsTab.append(lstCols)

  ## DEBUG - IMPRIME AS LINHAS EXTRAIDAS DA TABELA
  if printDebug == True:
    try:
      strFileName = "Saida/" + str(dtRefDados.year) + "_" + str(dtRefDados.month).zfill(2) + "_lista_Garantia_Fisica.txt"
      f = open(strFileName, "w")
      try:
        for i in xrange(0,len(lstRowsTab)):
          row = lstRowsTab[i]
          for i in xrange(0,len(row)):
            f.write(row[i] + ";")
          f.write("\n")
      finally:
        f.close()
    except IOError:
      pass

  ##Retorna a lista sem cabecalho
  return lstRowsTab[1:],dtRefDados
############################# FINAL DO getLstRows_TabGarantiaFisica_HtmlLocal





###############################################################################
#### METODO I - CONTRATOS DE COMPRA
## @autor PALIN
## Description: Extrai apenas a TABELA Contratos de Venda que esta entre
## strExpReg_TABConsumo = 'Perfil SiglaAgente(.*?)Tabela 2'
###############################################################################
def getLstRows_TabContratCompra_HtmlLocal(strConteudoHTML,printDebug = False):

  #Procura por uma data do tipo: 2013/03
  expRegData = re.compile(r'(([0-9]{4})/([0-9]{2}))')
  strDataRef = ''
  match = re.search(expRegData, strConteudoHTML)
  if match:
      strDataRef = match.group()

  #- SEPARANDO A DATA
  #Divide no / Armazenando o Mes e Ano
  strDataRef = str(strDataRef)
  (strAno,strMes) = strDataRef.split('/')

  ##DATA DE REFERENCIA DOS DADOS A SEREM INSERIDOS NO BD
  dtRefDados = datetime.datetime(int(strAno),int(strMes),1)

  ##1 PASSO
  ## A PAGINA DA CCEE EXISTE VARIAS TABELAS DENTRO DE TABELAS
  ## VAMOS PEGAR APENAS O HTML REFERENTE A TABELA QUE QUEREMOS USANDO
  ## EXTRACAO POR EXPRESSAO REGULAR:
  strExpReg_TAB = u'Mensal de Compra \(MWh\)(.*?)gina 1'
  strHtmlTab = ''
  for strBlocoFound in re.findall(strExpReg_TAB, strConteudoHTML, re.S):
    strHtmlTab = strBlocoFound

  ##(2) PASSO, dado que a expressao regular extraiu a partir da primeira
  idx0InicioTab2 = strHtmlTab.find("Perfil Agente")
  strTabFinal1 = strHtmlTab[idx0InicioTab2:]
  strInicioHtmlTab = "<table><tbody><tr><td> Perfil Agente"

  ##(3) Tabela reconstruida. Falta eliminar uma parte final apos encontrarmos
  ## o 1o </tbody></table>
  strTabFinal1 = strInicioHtmlTab + str(strTabFinal1)

  ##Encontrando o Fim
  idx0FinalTab = strTabFinal1.find("</tbody></table>")

  ##PRONTO - TENHO UM HTML APENAS COM A TABELA - VENDA
  ##Pegando + 17 caracteres para englobar a string </tbody></table>
  strTabFinalCompra = strTabFinal1[:idx0FinalTab+17]

  ##DEBUG - Escreve em Arquivo
  if printDebug == True:
    strFileName = "Saida/" + str(dtRefDados.year) + "_" + str(dtRefDados.month).zfill(2) + "_HTML_Contratos_Compra.html"
    try:
      f = open(strFileName, "w")
      try:
        f.writelines(strTabFinalCompra) # Escreve uma sequencia de Strings para o Arquivo
      finally:
        f.close()
    except IOError:
      pass

  ##(4) - Avaliando a Tabela (html) com BeautifulSoup
  soup = BeautifulSoup(strTabFinalCompra)

  ##(5) - BeautifulSoup encontra no Html quantas tabelas existem
  ##desejamos que ele encontre apenas 1 tabela
  lstTables = soup.find_all('table')
  print "Existem " + str(len(lstTables)) + " tabelas na pagina"

  ##Apos avaliarmos linha a linha, coluna a coluna
  lstRowsTab = []

  ## INICIO DA LEITURA - CADA LINHA LE TODAS AS COLUNAS
  rows = lstTables[0].findAll('tr')
  for tr in rows:
    cols = tr.findAll('td')
    lstCols = [] #Limpa a lista de Colunas para Proxima Linha
    for td in cols:
        strText = td.find(text=True)
        strText = strText.encode('latin-1').strip()
        ##Dado que o numero esta escrito: 1.454,56, apenas
        ##removemos o .(ponto)
        strText = strText.replace(".","")
        strText = strText.replace("\r","")
        strText = strText.replace("\n","")
        strText = strText.replace("\t","")
        ##Colocamos na lista o caracter Windows ANSI
        ##lstCols.append(strText.encode('cp1252')) ##Trata caracteres acentuados
        lstCols.append(strText) ##Trata caracteres acentuados
    lstRowsTab.append(lstCols)

    if printDebug == True:
      try:
        strFileName = "Saida/" + str(dtRefDados.year) + "_" + str(dtRefDados.month).zfill(2) + "_Lista_Contratos_Compra.txt"
        f = open(strFileName, "w")
        try:
          for i in xrange(0,len(lstRowsTab)):
            row = lstRowsTab[i]
            for i in xrange(0,len(row)):
              f.write(row[i] + ";")
            f.write("\n")
        finally:
          f.close()
      except IOError:
        pass

  ##Remove o cabecalho antes de devolver
  return lstRowsTab[1:],dtRefDados
############################# FINAL DO getLstRows_TabContratCompra_HtmlLocal




###############################################################################
#### METODO I - CONTRATOS DE VENDA
## @autor PALIN
## Description: Extrai apenas a TABELA Contratos de Venda que esta entre
## strExpReg_TABConsumo = 'Perfil SiglaAgente(.*?)Tabela 2'
## Problemas com acento: print(repr(b'J\xe2nis'.decode('cp1252')))
## Veja print(b'J\xe2nis'.decode('cp1252'))
###############################################################################
def getLstRows_TabContratVenda_HtmlLocal(strConteudoHTML,printDebug = False):

  #Procura por uma data do tipo: 2013/03
  expRegData = re.compile(r'(([0-9]{4})/([0-9]{2}))')
  strDataRef = ''
  match = re.search(expRegData, strConteudoHTML)
  if match:
      strDataRef = match.group()

  #- SEPARANDO A DATA
  #Divide no / Armazenando o Mes e Ano
  strDataRef = str(strDataRef)
  (strAno,strMes) = strDataRef.split('/')

  ##DATA DE REFERENCIA DOS DADOS A SEREM INSERIDOS NO BD
  dtRefDados = datetime.datetime(int(strAno),int(strMes),1)
  print " Data do Dados: " + dtRefDados.strftime('%m/%Y')

  strExpReg_TAB = 'Perfil Agente(.*?)Tabela 2'
  strHtmlTab = ''
  for strBlocoFound in re.findall(strExpReg_TAB, strConteudoHTML, re.S):
    strHtmlTab = strBlocoFound

  strInicioHtmlTab = "<table> <tbody> <tr> <td> Perfil Agente"

  ##(3) Tabela reconstruida. Falta eliminar uma parte final apos encontrarmos
  ## o 1o </tbody></table>
  strTabFinal1 = strInicioHtmlTab + str(strHtmlTab)
  idx0FinalTab = strTabFinal1.find("</tbody></table>")

  ##PRONTO - TENHO UM HTML APENAS COM A TABELA - VENDA
  ##Pegando + 17 caracteres para englobar a string </tbody></table>
  strTabFinalVenda = strTabFinal1[:idx0FinalTab+17]

  ##DEBUG - Escreve um TXT com o HTML extraido
  if printDebug == True:
    strFileName = "Saida/" + str(dtRefDados.year) + "_" + str(dtRefDados.month).zfill(2) + "_HTML_Contratos_Venda.html"
    try:
      f = open(strFileName, "w")
      try:
        f.writelines(strTabFinalVenda) # Escreve uma sequencia de Strings para o Arquivo
      finally:
        f.close()
    except IOError:
      pass

  ##(4) - Avaliando a Tabela (html) com BeautifulSoup
  soup = BeautifulSoup(strTabFinalVenda)

  ##(5) - BeautifulSoup encontra no Html quantas tabelas existem
  ##desejamos que ele encontre apenas 1 tabela
  lstTables = soup.find_all('table')
  print "Existem " + str(len(lstTables)) + " tabelas na pagina"

  ##Apos avaliarmos linha a linha, coluna a coluna
  lstRowsTab = []

  ## INICIO DA LEITURA - CADA LINHA LE TODAS AS COLUNAS
  rows = lstTables[0].findAll('tr')
  for tr in rows:
    cols = tr.findAll('td')
    lstCols = [] ##Limpa a lista de Colunas para Proxima Linha
    for td in cols:
        strText = td.find(text=True)
        strText = strText.encode('latin-1').strip()
        strText = strText.replace(".","")
        strText = strText.replace("\r","")
        strText = strText.replace("\n","")
        strText = strText.replace("\t","")
        lstCols.append(strText) ##Trata caracteres acentuados
    lstRowsTab.append(lstCols)

  ## DEBUG - IMPRIME AS LINHAS EXTRAIDAS DA TABELA
  if printDebug == True:
    try:
      strFileName = "Saida/" + str(dtRefDados.year) + "_" + str(dtRefDados.month).zfill(2) + "_Lista_Contratos_Venda.txt"
      f = open(strFileName, "w")
      try:
        for i in xrange(0,len(lstRowsTab)):
          row = lstRowsTab[i]
          for i in xrange(0,len(row)):
            f.write(row[i] + ";")
          f.write("\n")
      finally:
        f.close()
    except IOError:
      pass

  ##Retorna a lista sem o cabecalho
  return lstRowsTab[1:],dtRefDados
############################# FINAL DO getLstRows_TabContratVenda_HtmlLocal


##################################################################
## Leitura do Arquivo TXT,XML ou HTML -
## Armazena o Conteudo em uma String para posterior avaliacao
##################################################################
def readConteudoLocalFile(strFileName,strDirAndFileName):

  ##########################################
  ## MONTANDO O DIRETORIO BASE
  ## conforme SO - Linux ou Windows
  ##########################################
  if sys.platform == 'linux2' or sys.platform == 'linux':
      strDir = strDirAndFileName
  elif sys.platform == 'win32':
      strDir = strDirAndFileName

  #Armazenara o Conteudo do Arquivo (seja ele txt,xml ou html)
  strAllContentFile = ""
  try:
      # Tenta Ler o arquivo
      f = open(strDirAndFileName,"r")
      try:
        # Leitura do CONTEUDO TODO de uma Vez
        strAllContentFile = f.read()

        # (Opcao 2:) ou Leitura Linha a Linha
        #line = f.readline()
        ## (Opcao 3:) ou Leitura de Todas as Linhas em uma Lista
        #lstLines = f.readlines()

      finally:
          f.close()

  except IOError:
      pass

  return strAllContentFile


##########################################
## Leitura do Arquivo XML pegando uma
## lista de parametros iniciais da Rodada
##########################################
def readConteudoFile_vsII(strXMLFileName,strSubDir):

  ##########################################
  ## MONTANDO O DIRETORIO BASE
  ## conforme SO - Linux ou Windows
  ##########################################
  if sys.platform == 'linux2' or sys.platform == 'linux':
      strBaseDir = os.environ['HOME']
      strDir = strBaseDir + "/" + strSubDir + "/"
  elif sys.platform == 'win32':
      strBaseDir = os.getcwd()
      strDir = strBaseDir + "/" + strSubDir + "/"

  #Armazenara o Conteudo do XML
  strAllContentFile = ""
  try:
      # Tenta Ler o arquivo
      f = open(strDir + strXMLFileName,"r")
      try:
        # Leitura do CONTEUDO TODO de uma Vez
        strAllContentFile = f.read()

      finally:
          f.close()

  except IOError:
      pass

  return strAllContentFile








########################################################
## METODO : varre o diretorio dado (path)
## e retorna uma lista com todos os NOMES DE ARQUIVOS
## E TAMBEM (path+filename.extensao)
##    """
##    EXEMPLO de CHAMADA desse METODO
##    lstPathAndFiles = getList_NomeArq_E_Path(os.getcwd(),extensao)
##
##    ## Imprime um Log com os Diretorios
##    strFileNameLog =  "LOG_Arq_MCSDs_Encontrados.log"
##    fout = open(strFileNameLog, 'w')
##
##    for item in lstPathAndFiles:
##        fout.write(item[0] + "\n")
##        fout.write(item[1] + "\n")
##        #fout.write(item + "\n")
##    fout.close()
##    """
########################################################


def getList_NomeArq_E_Path(path,extensao):
    """
    Essa funcao retorna a lista de Arquivos e Caminho
    dos arquivos de uma dada extensao
    """
    lstFileName_DirAndFileName = []
    #conta o numero de arquivos na pasta
    for raiz, diretorios, arquivos in os.walk(os.getcwd()):
        for arquivo in arquivos:
            if arquivo.endswith(extensao):
                row = []
                ##Guarda o Nome do Arquivo
                row.append(arquivo)
                ##Nome do Arquivo e Diretorio
                row.append(os.path.join(raiz, arquivo))
                lstFileName_DirAndFileName.append(row)

    return lstFileName_DirAndFileName


########################################################
## METODO : varre o diretorio dado (path)
## e retorna uma lista com todos os NOMES DE ARQUIVOS
## E TAMBEM (path+filename.extensao)
##    """
##    EXEMPLO de CHAMADA desse METODO
##    lstPathAndFiles = getList_NomeArq_E_Path(os.getcwd(),extensao)
##
##    ## Imprime um Log com os Diretorios
##    strFileNameLog =  "LOG_Arq_MCSDs_Encontrados.log"
##    fout = open(strFileNameLog, 'w')
##
##    for item in lstPathAndFiles:
##        fout.write(item[0] + "\n")
##        fout.write(item[1] + "\n")
##        #fout.write(item + "\n")
##    fout.close()
##    """
########################################################

def getList_NomeArq_E_Path_2Ext(path,extensao1,extensao2):
    """
    Essa funcao retorna a lista de Arquivos e Caminho
    dos arquivos de uma dada extensao
    """
    lstFileName_DirAndFileName = []
    #para a extensao1
    for raiz, diretorios, arquivos in os.walk(os.getcwd()):
        for arquivo in arquivos:
            if arquivo.endswith(extensao1):
                row = []
                ##Guarda o Nome do Arquivo
                row.append(arquivo)
                ##Nome do Arquivo e Diretorio
                row.append(os.path.join(raiz, arquivo))
                lstFileName_DirAndFileName.append(row)

    #Para a extensao2
    for raiz, diretorios, arquivos in os.walk(os.getcwd()):
        for arquivo in arquivos:
            if arquivo.endswith(extensao2):
                row = []
                ##Guarda o Nome do Arquivo
                row.append(arquivo)
                ##Nome do Arquivo e Diretorio
                row.append(os.path.join(raiz, arquivo))
                lstFileName_DirAndFileName.append(row)

    return lstFileName_DirAndFileName



########################################################
## METODO 2: varre o diretorio dado (path)
## e retorna uma lista com todos os arquivos
## que inclui (path+filename.extensao)
## RETORNA UMA LISTA com CAMINHO e NOMES dos Arquivos
########################################################
def getListFilesWithDirs(path,extensao):
    """
    Essa funcao retorna a lista de Arquivos e Caminho
    dos arquivos de uma dada extensao
    """
    lstDirsAndFiles = []
    #conta o numero de arquivos na pasta
    for raiz, diretorios, arquivos in os.walk(path):
        for arquivo in arquivos:
            if arquivo.endswith(extensao):
                lstDirsAndFiles.append(os.path.join(raiz, arquivo))
                #os.remove(os.path.join(raiz, arquivo))
    return lstDirsAndFiles



########################################################
## METODO 3: imprime um LOG com os Dirs e Nomes
## dos arquivos .extensao encontrados nos Diretorios
########################################################
def printLogFilesFound(lstPathAndFiles,strFileNameLog):
    try:
        f = open(strFileNameLog, "w")
        try:
          for item in lstPathAndFiles:
              ##Escreve o Nome do Arquivo
              f.write(item[0] + "\n")
              ##Escreve o Diretorio + NomeDoArquivo
              f.write(item[1] + "\n")
        finally:
            f.close()
    except IOError:
        pass




def copyFile(src, dst):
    try:
        shutil.copytree(src, dst)
    # Depend what you need here to catch the problem
    except OSError as exc:
        # File already exist
        if exc.errno == errno.EEXIST:
            shutil.copy(src, dst)
        # The dirtory does not exist
        if exc.errno == errno.ENOENT:
            shutil.copy(src, dst)
        else:
            raise

##########################################
## METODO -
## No C++ map_MesAno_HistAfluenciaTodosSubSist
## Map entre Mes e Ana e o Vetor de Afluencias (2234,3434,6434,3433) = SE,S,NE,N

## DADO UMA TUPLA DO TIPO: (11, 2012, 'SE') ele retorna o Historico de Afluencia
## 23087.00
##(11, 2012, 'S') retorna 4877.00

##Exemplo de Uso:
  ## Dado a Tupla (12, 2012, 'SE')
  ## print dict_Data_HistAfluencia[(12, 2012, 'SE')]
##########################################
def getDic_MesAno_HistAfluencia(dtRef,dic_Subsist_LstEnas,lstCom11Tuplas_de_Datas):

  ##Retorna o Dicionario - dado uma Tupla:
  ## dict_Data_HistAfluencia[(12, 2012, 'SE')] fornece o Historico da Ena
  dict_Data_HistAfluencia = {}
  ##Faz uma Copia para NAO alterar a Data de Referencia
  dtAux = dtRef

  ## LISTA DE 11 MESES PARA TRAS DA DATA DE REFERENCIA
  lstDatas = []
  for i in xrange(1,12):
      dtAux = dtAux + relativedelta(months=-1)  #Mes Anterior
      lstDatas.append(dtAux)

  ##Coloca as Datas na Ordem da Menor para Maior para criar a Tuplas
  lstDatasReverse = list(reversed(lstDatas))
  ## IMPRIME a lista de DATAS na ordem CRESCENTE
  ##for i in xrange(0,len(lstDatasReverse)):
  ##    print lstDatasReverse[i].strftime('%d/%m/%Y')

  ## Criando a Lista de Tuplas:
  ##[(11, 2012, 'SE'), (11, 2012, 'S'), (11, 2012, 'NE'), (11, 2012, 'N')...
  lstSubm = ['SE','S','NE','N']

  lst4TuplasTemp = []

  for i in xrange(0,len(lstDatasReverse)):
      for k in xrange(0,len(lstSubm)):
          strSubM = lstSubm[k] ## SUBMERCADO
          idMes = lstDatasReverse[i].month ## MES
          idAno = lstDatasReverse[i].year ## ANO
          ## MONTA O DICIONARIO (11, 2012, 'SE') = dic_Subsist_LstEnas['SE'][11-1]
          ## Em dic_Subsist_LstEnas temos os valores nas Posicoes 0-index, por isso idMes-1
          dict_Data_HistAfluencia[(idMes, idAno, strSubM) ] = dic_Subsist_LstEnas[strSubM][idMes-1]
          lst4TuplasTemp.append((lstDatasReverse[i].month, lstDatasReverse[i].year,strSubM))


      ## Retorna tambem a Lista de Tuplas possiveis, combinando as 11 datas com os 4
      ## submercados
      lstCom11Tuplas_de_Datas.append(lst4TuplasTemp)
      lst4TuplasTemp = []

  return dict_Data_HistAfluencia




##########################################
## Leitura do Arquivo TXT
## Faz a leitura para uma Lista de Linhas
## que e' retornada no final
##########################################
def readFile_getListaLinhas(strFileName,strSubDir):

  ##########################################
  ## MONTANDO O DIRETORIO BASE
  ## conforme SO - Linux ou Windows
  ##########################################
  if sys.platform == 'linux2' or sys.platform == 'linux':
      strBaseDir = os.environ['HOME']
      strDir = strBaseDir + "/" + strSubDir + "/"
  elif sys.platform == 'win32':
      strBaseDir = os.getcwd()
      strDir = strBaseDir + "/" + strSubDir + "/"

  #Armazenara o Conteudo do TXT
  lstLines = []
  try:
      # Tenta Ler o arquivo
      f = open(strDir + strFileName,"r")
      try:
        # (Opcao 1:) Leitura do CONTEUDO TODO de uma Vez
        #strAllContentFile = f.read()

        # (Opcao 2:) ou Leitura Linha a Linha
        #line = f.readline()

        ## (Opcao 3:) ou Leitura de Todas as Linhas em uma Lista
        lstLines = f.readlines()

      finally:
          f.close()

  except IOError:
      pass

  return lstLines




## Tutorial http://stackoverflow.com/questions/5093002/finding-elements-by-attribute-with-lxml
##########################################
## Interpreta o XML passando a lista
## de Parametros Lida
## Pacotes: import StringIO  e lxml
## Parametros especificos do Projeto
## Relatorio de Inteligencia de Mercado
##########################################
#http://lxml.de/xpathxslt.html

def getParamFromXML_RIM(xmldoc):
  tree = etree.parse(StringIO(xmldoc))

  param_dict = {}
  elem = tree.xpath('/ConfiguracoesEntrada/Ano')
  #print "Tam: " + str(len(elem))
  #print elem[0].tag   #Imprime o nome do campo: "Ano"
  #print elem[0].text   #Imprime Conteudo de Ano
  #print elem[0].tag + " => " + elem[0].text
  param_dict[elem[0].tag] = elem[0].text

  #[1]
  elem = tree.xpath('/ConfiguracoesEntrada/Mes')
  #print elem[0].tag + " => " + elem[0].text
  param_dict[elem[0].tag] = elem[0].text

  #[2] DiretorioEntradaDados
  elem = tree.xpath('/ConfiguracoesEntrada/DiretorioEntradaDados')
  #print elem[0].tag + " => " + elem[0].text
  param_dict[elem[0].tag] = elem[0].text

  #[3] DiretorioSaidaDados
  elem = tree.xpath('/ConfiguracoesEntrada/DiretorioSaidaDados')
  #print elem[0].tag + " => " + elem[0].text
  param_dict[elem[0].tag] = elem[0].text

  return param_dict


#### METODO IV ####
## @autor PALIN
## Description: Baixa o arquivo Zip para o Diretorio Definido
def downloadFile(urlFile, fileName, pathDirDownload, auth_proxy=True):
    print "Baixando o Arquivo %s" % (fileName)

    strConexaoProxy = ''
    if auth_proxy == True:
        if sys.platform == 'linux2' or sys.platform == 'linux':
            strHomeDir = os.environ['HOME'] #diretorio Home do User
            #strHomeDir = "/root" #diretorio Home do User
            fileConexaoProxy = strHomeDir + "/.tmp/DadosProxy.txt"
        else:
            strHomeDir = 'c:/dev'
            fileConexaoProxy = strHomeDir + "/DadosProxy.txt"
        try:
            f = open(fileConexaoProxy)
        except:
            print "Erro 1 na Leitura do Arquivo: " + str(fileConexaoProxy)
            sys.exit()

        strConexaoProxy = f.readline()
        f.close()
        #++++ FINAL CONEXAO COM PROXY - EDP ++++++
    if strConexaoProxy != '':
        strConexaoProxy = strConexaoProxy.strip().split()[0]

    http_proxy  = strConexaoProxy
    https_proxy = strConexaoProxy
    ftp_proxy   = strConexaoProxy
    proxyDict = {
                  "http"  : http_proxy,
                  "https" : https_proxy,
                  "ftp"   : ftp_proxy
                }


    if auth_proxy == True:
      r = requests.get(urlFile,proxies=proxyDict)
    else:
      r = requests.get(urlFile)

    #print r.headers.get('content-type')
    #print r.headers
    #print r.headers.get('content-length')
    file_size = int(r.headers.get('content-length'))
    print "Baixando: %s %s Kb" % (fileName, file_size/1024)

    #Muda para o diretorio onde deve baixar
    os.chdir(pathDirDownload)
    # Writes the data to a local file one chunk at a time.
    f = open(fileName, 'wb')
    for chunk in r.iter_content(chunk_size = 512 * 1024): # Reads 512KB at a time into memory
        if chunk: # filter out keep-alive new chunks
            f.write(chunk)
    f.close()


#### METODO II ####
## @autor PALIN
## Description:
#JUNHO 2013 - autenticacao Proxy agora funciona no Windows tambem
#Problema era a string de conexao onde deveriamos ter
#http://escelsa\ec204305:XXXX@band24:8080  uma barra apos escelsa
#http://monzool.net/blog/2007/10/15/html-parsing-with-beautiful-soup/
def downloadFileUrlib(urlFile, fileName, pathDirDownload, auth_proxy=True):
    strConexaoProxy = ''
    if auth_proxy == True:
        if sys.platform == 'linux2' or sys.platform == 'linux':
            strHomeDir = os.environ['HOME'] #diretorio Home do User
            #strHomeDir = "/root" #diretorio Home do User
            fileConexaoProxy = strHomeDir + "/.tmp/DadosProxy.txt"
        else:
            strHomeDir = 'c:/dev'
            fileConexaoProxy = strHomeDir + "/DadosProxy.txt"
        try:
            f = open(fileConexaoProxy)
        except:
            print "Erro 4 na Leitura do Arquivo: " + str(fileConexaoProxy)
            sys.exit()

        strConexaoProxy = f.readline()
        f.close()
        #++++ FINAL CONEXAO COM PROXY - EDP ++++++

    #print "Usando string: " + str(strConexaoProxy)
    #REMOVE ESPACO E ENTER DO FINAL DA STRING
    if strConexaoProxy != '':
        strConexaoProxy = strConexaoProxy.strip().split()[0]

    strConexao = "[CONEXAO_PROXY]"
    strConexao = strConexao.replace("[CONEXAO_PROXY]",strConexaoProxy)
    proxy_support = urllib2.ProxyHandler({"http":strConexao})
    opener = urllib2.build_opener(proxy_support)

    if auth_proxy == True:
        urllib2.install_opener(opener)

    ## leitura do conteudo WEB
    u = ""
    try:
        u = urllib2.urlopen(urlFile)
    except urllib2.HTTPError:
        print "Link: " + urlFile + " indisponivel!"
        sys.exit()


    try:
        #Muda para o diretorio onde deve baixar
        os.chdir(pathDirDownload)
        f = open(fileName, 'wb')
        meta = u.info()
        #print "Imprindo as Informacoes da Pagina: " + str(meta)
        #info() fornece varis informacoes sobre o arquivo - vamos pegar o Tamanho
        file_size = int(meta.getheaders("Content-Length")[0])
        print "Baixando: %s %s Mb" % (fileName, file_size/1024)

        file_size_dl = 0
        block_sz = 8192
        while True:
            buffer = u.read(block_sz)
            if not buffer:
                break

            file_size_dl += block_sz
            f.write(buffer)
            status = r"%10d  [%3.2f%%]" % (file_size_dl, file_size_dl * 100. / file_size)
            status = status + chr(8)*(len(status)+1)
            print status,

        f.close()
        return 0 # Retirado para podermos continuar ate o final do codigo

    except IOError:
        f.close()
        return 1 # Falhou o Download


## USO: zipDirToFile('c:/test', 'c:/temp/test.zip')

#Zipa Diretorio e Subdiretorios
def zipDirToFile(dir, zip_file):
    zip = zipfile.ZipFile(zip_file, 'w', compression=zipfile.ZIP_DEFLATED)
    root_len = len(os.path.abspath(dir))
    for root, dirs, files in os.walk(dir):
        archive_root = os.path.abspath(root)[root_len:]
        for f in files:
            fullpath = os.path.join(root, f)
            archive_name = os.path.join(archive_root, f)
            #print f
            zip.write(fullpath, archive_name, zipfile.ZIP_DEFLATED)
    zip.close()
    return zip_file


def zipUmArquivo(pathAndNameFileToZip, new_zip_file):
    print 'Criando o arquivo: ' + new_zip_file
    zf = zipfile.ZipFile(new_zip_file, mode='w')
    try:
        zf.write(pathAndNameFileToZip)
    finally:
        print 'Finalizando ' + new_zip_file
        zf.close()


def getZipContent(dir, zip_file):
  zf = zipfile.ZipFile(zip_file, 'r')
  return zf.namelist()


## RETORNA TRUE OR FALSE
def hasProxy():
  print "### TENTANDO IDENTIFICAR SE PRECISA O USO DE PROXY OU NAO ###"
  ### INICIO DA TENTATIVA DE DETECTAR PROXY AUTOMATICAMENTE
  strConexaoProxy = ''
  if sys.platform == 'linux2' or sys.platform == 'linux':
      strHomeDir = os.environ['HOME'] #diretorio Home do User
      #strHomeDir = "/root" #diretorio Home do User
      fileConexaoProxy = strHomeDir + "/.tmp/DadosProxy.txt"
  else:
      strHomeDir = 'c:/dev'
      fileConexaoProxy = strHomeDir + "/DadosProxy.txt"
  try:
      f = open(fileConexaoProxy)
  except:
      print "Erro 1 na Leitura do Arquivo: " + str(fileConexaoProxy)
      sys.exit()

  strConexaoProxy = f.readline()
  print "Lido conexao proxy: " + strConexaoProxy
  f.close()
  #++++ FINAL DA LEITURA DE CONEXAO COM PROXY - EDP ++++++

  if strConexaoProxy != '':
      strConexaoProxy = strConexaoProxy.strip().split()[0]

  http_proxy  = strConexaoProxy
  https_proxy = strConexaoProxy
  ftp_proxy   = strConexaoProxy
  proxyDict = {
                "http"  : http_proxy,
                "https" : https_proxy,
                "ftp"   : ftp_proxy
              }

  ## 1 - tenta acessar algum conteudo WEB

  ## Admite inicialmente que usa proxy
  flgUsaProxy = True


  urlCheck = "http://www.google.com.br"
  try:
    r = requests.get(urlCheck,proxies=proxyDict)
  except requests.exceptions.Timeout:
    # Maybe set up for a retry, or continue in a retry loop
    print "Erro 1"
    print e # catastrophic error. bail.
    sys.exit(1)
  except requests.exceptions.TooManyRedirects:
    # Tell the user their URL was bad and try a different one
    print "Erro 2"
    print e # catastrophic error. bail.
    sys.exit(1)
  except requests.exceptions.RequestException as e:
    print "Erro 3"
    print e # catastrophic error. bail.
    try:
      r = requests.get(urlCheck)
      flgUsaProxy = False
    except requests.exceptions.RequestException as e:
      print "Tentei novamente e nao consegui:"
      sys.exit(1)
  else:
      print "Conexao OK - Nao Usa Proxy"
      flgUsaProxy = True

  if flgUsaProxy == True:
    print "DETECTEI PROXY"
  elif flgUsaProxy == False:
    print "NAO TEM PROXY"

  return flgUsaProxy


##########################################
## Retorna a Lista dos Meses do Ano
## Capitalizada (1a Maiuscula) em Pt-br
##########################################
def getNomesMeses():
  ## PARA RESOLVER PROBLEMAS COM ACENTUACAO NAO ESQUECA DE:
  #IMPORTAR O PACKAGE "string"
  #PARA APRESENTAR ALGUM NOME COM ACENTO NA TELA DEVEMOS FAZER:
  #print strNomeAcentuado.encode('latin-1')
  #Imprime a Data em Portugues
  try:
      setlocale(LC_ALL,'portuguese')
  except:
      setlocale(LC_ALL,'pt_BR.UTF-8')

  #Pegando os Nomes dos Meses
  #lstMeses = [mes for mes in calendar.month_name]

  #Pegando os Meses com a Primeira Letra Maiuscula (capitalize)
  lstMeses = [mes.capitalize() for mes in calendar.month_name if mes]

  ## IMPRIME LISTA
  #for i in xrange(0, len(lstMeses)):
  #  print lstMeses[i]

  return lstMeses


##########################################
## Metodo Verifica se a String que foi
## passada e' um Numero Valido
##########################################
def is_number(strTxt):
    try:
        float(strTxt) # for int, long and float
    except ValueError:
        try:
            complex(strTxt) # for complex
        except ValueError:
            return False

    return True


##########################################
## Metodo Verifica se a String que foi
## passada e' um Numero Valido
##########################################
def info_System():
  import sys
  print "++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"
  print "-- INTELIGENCIA MERCADO TO BD --"

  #Imprime a Data em Portugues
  try:
      setlocale(LC_ALL,'portuguese')
  except:
      setlocale(LC_ALL,'pt_BR.UTF-8')

  ##(1) PARA RESOLVER PROBLEMAS COM ACENTUACAO NAO ESQUECA DE:
  #IMPORTAR O PACKAGE "string"

  #(2) PARA APRESENTAR ALGUM NOME COM ACENTO NA TELA DEVEMOS FAZER:
  #print strNomeAcentuado.encode('latin-1')
  print locale.getlocale() #tem que imprimir ('pt_BR', 'cp1252')
  now = datetime.datetime.now()
  print "Inicio Rodada: " + now.strftime('%A %d de %B de %Y, %H:%M:%S')
  print "Nome do SO: " + string.upper(os.name) + " Plataforma: " + string.upper(sys.platform)
  print "Diretorio Local: " + os.getcwd()
  print "++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"


##########################################
## Metodo Retorna o Nome Extenso do Mes
## atual
##########################################
def getNomeMes(idMes):

    #Packages Necessarios
    #from locale import setlocale, LC_ALL
    #import calendar

    try:
        setlocale(LC_ALL,'portuguese')
    except:
        setlocale(LC_ALL,'pt_BR.UTF-8')

    return calendar.month_name[int(idMes)]



##########################################
## Metodo Escreve um Conteudo TXT
## para um arquivo.
## Param: NomeArq e Conteudo
##########################################
def writeFile(fileName,strConteudo):
    try:
        ## This will create a new file or **overwrite an existing file**.
        f = open(fileName, "w")
        try:
            ##f.write("word") # Escreve a string apenas
            f.write(strConteudo.encode('utf-8'))
            ##f.write(strConteudo)
        finally:
            f.close()
    except IOError:
        pass
##------------------------------------------------------------------------------



# getNomeMes usa o locale e calendar ###
#    print '# Teste com formatação:',  formata_ToMoeda(10.50)
def getNomeMes(idMes):
    from locale import setlocale, LC_ALL
    import calendar
    try:
        setlocale(LC_ALL,'portuguese')
    except:
        setlocale(LC_ALL,'pt_BR.UTF-8')

    return calendar.month_name[int(idMes)]


# getNomeMes usa o locale e calendar ###
#    print '# Teste com formatação:',  formata_ToMoeda(10.50)
def getNomeMesAbreviado(idMes):
    from locale import setlocale, LC_ALL
    import calendar
    try:
        setlocale(LC_ALL,'portuguese')
    except:
        setlocale(LC_ALL,'pt_BR.UTF-8')

    return string.upper(calendar.month_abbr[int(idMes)])





##########################################
## Define qual e' o Metodo Principal
## que sera chamado ao executar o Script
## Para alterar - altere a linha main()
##########################################
if __name__ == '__main__':
    main()